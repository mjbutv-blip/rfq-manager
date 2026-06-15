"""
生成全流程测试用询单表 Excel

数据覆盖：
  - 11 条正常行（不同业务员/小组/客户/产品大类/季节/状态）
  - 3 条 failed 行（缺少必填字段）
  - 1 条文件内重复行（duplicate）
  - 1 条空行（解析器跳过）
  共 16 条数据行

用法：
  cd backend
  python scripts/create_test_excel.py              # 输出到 scripts/test_inquiry.xlsx
  python scripts/create_test_excel.py /path/out.xlsx
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    "询单号", "客户代码", "客户简称", "国家/地区", "地区",
    "所属小组", "负责业务员", "协助业务员",
    "产品大类", "品名", "系列名称", "季节",
    "下单数量", "询单日期", "报价情况", "订单状态",
    "最终报价", "工厂价格", "毛利率", "下单单价", "数量", "贸易额",
    "下单日期", "备注",
]

# fmt: off
ROWS = [
    # ── 正常行 × 11 ──────────────────────────────────────────────────────────────
    # A组 张伟 泳装 2026Q1 已下单 有完整财务数据
    ["MVT-001", "C001", "泰格体育",   "美国",   "北美",
     "A组", "张伟", "李娜",
     "泳装", "男童泳裤",       "SS2026泳装系列", "2026春夏",
     500, date(2026, 1, 10), "已报价", "下单",
     12.50, 8.00, "18.5%", 12.50, 500, 6250.00, date(2026, 2, 1),  "加急"],

    # A组 张伟 泳装 2026Q1 跟进中 有报价无下单
    ["MVT-002", "C001", "泰格体育",   "美国",   "北美",
     "A组", "张伟", "",
     "泳装", "女童泳衣",       "SS2026泳装系列", "2026春夏",
     300, date(2026, 1, 15), "已报价", "跟进中",
     15.00, 9.50, "22.5%", 0, 0, 0, None, ""],

    # A组 王芳 运动服 2026Q2 已下单
    ["MVT-003", "C002", "维多利亚",   "英国",   "欧洲",
     "A组", "王芳", "",
     "运动服", "女款运动套装",  "2026运动系列",   "2026秋冬",
     200, date(2026, 4, 2),  "已报价", "下单",
     45.00, 28.00, "37.8%", 45.00, 200, 9000.00, date(2026, 5, 10), ""],

    # A组 王芳 运动服 2026Q2 报价中
    ["MVT-004", "C002", "维多利亚",   "英国",   "欧洲",
     "A组", "王芳", "",
     "运动服", "男款运动背心",  "2026运动系列",   "2026秋冬",
     150, date(2026, 4, 20), "报价中", "跟进中",
     0, 0, "", 0, 0, 0, None, ""],

    # B组 李梅 泳装 2026Q2 已下单
    ["MVT-005", "C003", "松本贸易",   "日本",   "亚太",
     "B组", "李梅", "",
     "泳装", "连体泳衣",       "SS2026泳装系列", "2026春夏",
     400, date(2026, 4, 5),  "已报价", "下单",
     22.00, 14.00, "25.0%", 22.00, 400, 8800.00, date(2026, 5, 15), ""],

    # B组 李梅 内衣 2026Q2 跟进中
    ["MVT-006", "C003", "松本贸易",   "日本",   "亚太",
     "B组", "李梅", "",
     "内衣", "女款内衣套装",   "内衣系列",       "2026秋冬",
     600, date(2026, 5, 1),  "询价中", "跟进中",
     0, 0, "", 0, 0, 0, None, ""],

    # B组 赵磊 户外 2025Q3 已下单
    ["MVT-007", "C004", "山地户外",   "德国",   "欧洲",
     "B组", "赵磊", "",
     "户外", "防风冲锋衣",     "户外系列",       "2025秋冬",
     80, date(2025, 9, 10),  "已报价", "下单",
     120.00, 75.00, "15.0%", 120.00, 80, 9600.00, date(2025, 10, 20), ""],

    # B组 赵磊 运动服 2025Q4 流失
    ["MVT-008", "C004", "山地户外",   "德国",   "欧洲",
     "B组", "赵磊", "",
     "运动服", "登山速干T恤",  "户外系列",       "2025秋冬",
     200, date(2025, 10, 1), "已报价", "流失",
     35.00, 22.00, "37.1%", 0, 0, 0, None, ""],

    # A组 张伟 内衣 2025Q4 已下单 大单
    ["MVT-009", "C001", "泰格体育",   "美国",   "北美",
     "A组", "张伟", "李娜",
     "内衣", "男款内裤套装",   "内衣系列",       "2025秋冬",
     1000, date(2025, 11, 15), "已报价", "下单",
     8.50, 5.50, "30.0%", 8.50, 1000, 8500.00, date(2026, 1, 5), "大单"],

    # B组 李梅 泳装 2025Q1 已下单
    ["MVT-010", "C003", "松本贸易",   "日本",   "亚太",
     "B组", "李梅", "",
     "泳装", "儿童泳帽",       "SS2025泳装系列", "2025春夏",
     800, date(2025, 3, 20), "已报价", "下单",
     5.50, 3.50, "22.5%", 5.50, 800, 4400.00, date(2025, 4, 25), ""],

    # A组 王芳 运动服 2025Q4 跟进中
    ["MVT-011", "C002", "维多利亚",   "英国",   "欧洲",
     "A组", "王芳", "",
     "运动服", "女童运动外套",  "2025运动系列",   "2025秋冬",
     300, date(2025, 10, 10), "已报价", "跟进中",
     28.00, 18.00, "35.7%", 0, 0, 0, None, ""],

    # ── failed 行 × 3 ────────────────────────────────────────────────────────────
    # 缺 group_name
    ["MVT-012", "C004", "山地户外",   "德国",   "欧洲",
     "",  "赵磊", "",               # group_name 为空
     "户外", "防水背包",     "户外系列",       "2026秋冬",
     50, date(2026, 3, 1),  "", "",
     0, 0, "", 0, 0, 0, None, "缺group_name"],

    # 缺 product_name
    ["MVT-013", "C001", "泰格体育",   "美国",   "北美",
     "A组", "张伟", "",
     "泳装", "",                    # product_name 为空
     "SS2026泳装系列", "2026春夏",
     100, date(2026, 2, 10), "", "",
     0, 0, "", 0, 0, 0, None, "缺product_name"],

    # 缺 quantity（数量列为空）
    ["MVT-014", "C002", "维多利亚",   "英国",   "欧洲",
     "A组", "王芳", "",
     "运动服", "男款冲锋衣",  "2026运动系列",   "2026秋冬",
     None, date(2026, 3, 5), "", "",   # order_quantity(下单数量)=None, OK
     0, 0, "", 0, None, 0, None, "缺quantity"],  # quantity(数量)=None → failed

    # ── duplicate 行（inquiry_no 与 MVT-002 重复）────────────────────────────────
    ["MVT-002", "C001", "泰格体育(重复)", "美国", "北美",
     "A组", "张伟", "",
     "泳装", "女童泳衣(重复)", "SS2026泳装系列", "2026春夏",
     300, date(2026, 1, 15), "已报价", "",
     0, 0, "", 0, 0, 0, None, "文件内重复行"],

    # ── 空行（解析器应跳过）──────────────────────────────────────────────────────
    [""] * 24,
]
# fmt: on


def create_test_excel(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "询单总表"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 32)

    wb.save(output_path)

    normal = sum(1 for r in ROWS if r[0] and r[0] not in ("MVT-002", "") and not any(
        str(c) in ("缺group_name", "缺product_name", "缺quantity") for c in r))
    print(f"✓ 测试 Excel 已生成：{output_path}")
    print(f"  总数据行：{len(ROWS)} 行")
    print(f"  正常行 11 条（3 小组 / 3 业务员 / 4 客户 / 4 产品大类 / 2026&2025 跨年）")
    print(f"  failed 行 3 条（缺 group_name / product_name / quantity）")
    print(f"  duplicate 行 1 条（inquiry_no=MVT-002 重复）")
    print(f"  空行 1 条（跳过）")


if __name__ == "__main__":
    default = Path(__file__).parent / "test_inquiry.xlsx"
    output = sys.argv[1] if len(sys.argv) > 1 else str(default)
    create_test_excel(output)
