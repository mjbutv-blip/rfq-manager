"""
数据备份路由

POST /backups/generate            — admin 生成系统备份
GET  /backups                     — admin 查看备份历史
GET  /backups/{backup_id}/download — admin 下载备份文件
POST /backups/restore/preview     — admin 上传备份文件预览
"""
from __future__ import annotations

import json
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import FileResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.permissions import UserDep
from app.database import get_db
from app.models.backup_record import BackupRecord
from app.models.operation_log import OperationLog

router = APIRouter(prefix="/backups", tags=["backups"])
DbDep = Annotated[AsyncSession, Depends(get_db)]

BACKUP_DIR = Path(__file__).parent.parent.parent / "generated" / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

# 备份表配置：(表名, 导出列 None=全部)
BACKUP_TABLES: list[tuple[str, list[str] | None]] = [
    ("inquiries",            None),
    ("inquiry_items",        None),
    ("customers",            None),
    ("factories",            None),
    ("factory_quote_records", None),
    ("sample_records",       None),
    ("production_records",   None),
    ("inquiry_warnings",     None),
    ("transfer_orders",      None),
    ("import_batches",       None),
    ("import_rows",          None),
    ("operation_logs",       None),
    ("groups",               None),
    # users 排除密码
    ("users", [
        "id", "username", "display_name", "role", "group_name",
        "email", "is_active", "is_pending", "last_login_at",
        "created_at", "updated_at",
    ]),
]

# 各表期望的必填字段（用于恢复预览校验）
REQUIRED_COLUMNS: dict[str, list[str]] = {
    "inquiries":     ["inquiry_no", "responsible_sales", "group_name"],
    "customers":     ["customer_code"],
    "factories":     ["factory_code"],
    "users":         ["username", "role"],
}


# ── 辅助：写操作日志 ───────────────────────────────────────────────────────────

async def _log(
    db: AsyncSession,
    actor: object,
    action_type: str,
    description: str,
    after_data: dict | None = None,
    request: Request | None = None,
) -> None:
    db.add(OperationLog(
        id=uuid.uuid4(),
        actor_username=actor.username,      # type: ignore[attr-defined]
        actor_display_name=actor.display_name,  # type: ignore[attr-defined]
        actor_role=actor.role,              # type: ignore[attr-defined]
        action_type=action_type,
        target_type="backup",
        description=description,
        after_data_json=after_data,
        status="success",
        request_path=str(request.url.path) if request else None,
        request_method=request.method if request else None,
        ip_address=request.client.host if request and request.client else None,
    ))


# ── 辅助：从 DB 读取表数据 ─────────────────────────────────────────────────────

async def _fetch_table(db: AsyncSession, table: str, columns: list[str] | None) -> tuple[list[str], list[list]]:
    """返回 (headers, rows)，rows 中每项是列表。"""
    col_sql = ", ".join(f'"{c}"' for c in columns) if columns else "*"
    try:
        result = await db.execute(text(f'SELECT {col_sql} FROM "{table}"'))
        keys = list(result.keys())
        rows = [list(row) for row in result.fetchall()]
        return keys, rows
    except Exception:
        return columns or [], []


# ── 辅助：构建 Excel workbook ─────────────────────────────────────────────────

def _build_excel(table_data: dict[str, tuple[list[str], list[list]]]) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)   # 删除默认空 sheet

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center")

    for sheet_name, (headers, rows) in table_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])   # Excel sheet 名最长 31 字符

        # 表头
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # 数据行
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                if isinstance(val, datetime):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val.replace(tzinfo=None))
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"
                elif isinstance(val, (dict, list)):
                    ws.cell(row=row_idx, column=col_idx, value=json.dumps(val, ensure_ascii=False))
                elif isinstance(val, uuid.UUID):
                    ws.cell(row=row_idx, column=col_idx, value=str(val))
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)

        # 冻结首行、自动筛选
        ws.freeze_panes = "A2"
        if headers:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # 自动列宽
        for col_idx, h in enumerate(headers, 1):
            col_vals = [str(h)] + [str(rows[r][col_idx - 1] or "") for r in range(min(50, len(rows)))]
            max_len  = max((len(v) for v in col_vals), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 1. 生成备份 ───────────────────────────────────────────────────────────────

@router.post("/generate")
async def generate_backup(db: DbDep, user: UserDep, request: Request):
    if user.role != "admin":
        raise HTTPException(403, "仅管理员可生成备份")

    now     = datetime.now(tz=timezone.utc)
    ts      = now.strftime("%Y%m%d_%H%M%S")
    fname   = f"系统备份_{ts}.xlsx"
    fpath   = BACKUP_DIR / fname

    record = BackupRecord(
        id=uuid.uuid4(),
        backup_name=f"系统备份 {now.strftime('%Y-%m-%d %H:%M:%S')}",
        backup_type="manual_excel",
        file_name=fname,
        generated_by=user.username,
        generated_at=now,
        status="failed",
    )
    db.add(record)
    await db.flush()

    try:
        table_data: dict[str, tuple[list[str], list[list]]] = {}
        row_counts: dict[str, int] = {}
        included: list[str] = []

        for table_name, columns in BACKUP_TABLES:
            headers, rows = await _fetch_table(db, table_name, columns)
            table_data[table_name] = (headers, rows)
            row_counts[table_name] = len(rows)
            included.append(table_name)

        xlsx_bytes = _build_excel(table_data)
        fpath.write_bytes(xlsx_bytes)

        record.status              = "generated"
        record.file_path           = str(fpath)
        record.file_size           = len(xlsx_bytes)
        record.included_tables_json = included
        record.row_counts_json     = row_counts

        total_rows = sum(row_counts.values())
        await _log(db, user, "backup_generate",
                   f"生成备份：{fname}，共 {len(included)} 张表 {total_rows} 行",
                   {"file_name": fname, "file_size": len(xlsx_bytes),
                    "row_counts": row_counts, "tables": included},
                   request)
        await db.commit()

    except Exception as exc:
        record.status        = "failed"
        record.error_message = str(exc)
        await db.commit()
        raise HTTPException(500, f"备份生成失败：{exc}") from exc

    return {
        "backup_id":       str(record.id),
        "file_name":       fname,
        "download_url":    f"/api/v1/backups/{record.id}/download",
        "included_tables": included,
        "row_counts":      row_counts,
        "file_size":       len(xlsx_bytes),
        "message":         "备份生成成功",
    }


# ── 2. 备份历史 ───────────────────────────────────────────────────────────────

@router.get("")
async def list_backups(
    db: DbDep,
    user: UserDep,
    status: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    if user.role != "admin":
        raise HTTPException(403, "仅管理员可查看备份历史")

    q = select(BackupRecord).order_by(BackupRecord.created_at.desc())
    if status:
        q = q.where(BackupRecord.status == status)

    result = await db.execute(q.offset((page - 1) * page_size).limit(page_size))
    records = result.scalars().all()

    def _fmt(r: BackupRecord) -> dict:
        return {
            "backup_id":       str(r.id),
            "backup_name":     r.backup_name,
            "file_name":       r.file_name,
            "file_size":       r.file_size,
            "generated_by":    r.generated_by,
            "generated_at":    r.generated_at.isoformat() if r.generated_at else None,
            "status":          r.status,
            "included_tables": r.included_tables_json or [],
            "row_counts":      r.row_counts_json or {},
            "error_message":   r.error_message,
        }

    return [_fmt(r) for r in records]


# ── 3. 下载备份 ───────────────────────────────────────────────────────────────

@router.get("/{backup_id}/download")
async def download_backup(backup_id: str, db: DbDep, user: UserDep, request: Request):
    if user.role != "admin":
        raise HTTPException(403, "仅管理员可下载备份")

    result = await db.execute(
        select(BackupRecord).where(BackupRecord.id == uuid.UUID(backup_id))
    )
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(404, "备份记录不存在")
    if not record.file_path or not Path(record.file_path).exists():
        raise HTTPException(404, "备份文件不存在，可能已被删除")

    await _log(db, user, "backup_download",
               f"下载备份：{record.file_name}",
               {"backup_id": backup_id, "file_name": record.file_name},
               request)
    await db.commit()

    return FileResponse(
        path=record.file_path,
        filename=record.file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── 4. 恢复预览 ───────────────────────────────────────────────────────────────

@router.post("/restore/preview")
async def restore_preview(
    db: DbDep,
    user: UserDep,
    request: Request,
    file: UploadFile = File(...),
):
    if user.role != "admin":
        raise HTTPException(403, "仅管理员可使用恢复预览")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 格式的备份文件")

    content = await file.read()
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(400, "文件过大（最大 100MB）")

    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"无法解析文件：{exc}") from exc

    sheets_info: list[dict] = []
    can_restore = True

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = iter(ws.iter_rows(values_only=True))
        header_row = next(rows_iter, None)
        headers = [str(c) for c in (header_row or []) if c is not None]
        row_count = sum(1 for _ in rows_iter)

        required = REQUIRED_COLUMNS.get(sheet_name, [])
        missing  = [c for c in required if c not in headers]
        status   = "warning" if missing else "ok"
        if missing:
            can_restore = False

        sheets_info.append({
            "sheet_name":      sheet_name,
            "row_count":       row_count,
            "status":          status,
            "missing_columns": missing,
        })

    wb.close()

    total_rows = sum(s["row_count"] for s in sheets_info)
    await _log(db, user, "backup_restore_preview",
               f"恢复预览：{file.filename}，{len(sheets_info)} 个 sheet，{total_rows} 行",
               {"file_name": file.filename,
                "sheets": sheets_info,
                "can_restore": can_restore},
               request)
    await db.commit()

    return {
        "file_name":   file.filename,
        "sheets":      sheets_info,
        "can_restore": can_restore,
        "message":     "备份文件可识别。第一版仅支持预览，不会写入数据库。",
    }
