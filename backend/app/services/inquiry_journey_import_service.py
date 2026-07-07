from __future__ import annotations

import io
import json
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app import crud
from app.core.permissions import can_edit_inquiry
from app.models import Factory, FactoryQuoteRecord, Inquiry, OperationLog, QuoteItem
from app.services.operation_log_service import log_kwargs_from_user, safe_log

PRICE_UNIT = "件"
DOMESTIC = "domestic"
OVERSEAS = "overseas"


INQUIRY_FIELD_LABELS = {
    "customer_order_no": "客户订单号",
    "group_name": "所属小组",
    "responsible_sales": "负责业务员",
    "season": "季节",
    "product_name": "品名",
    "series_name": "系列",
    "quantity": "数量",
    "order_status": "订单状态",
    "order_date": "下单日期",
    "order_unit_price": "下单美金价格",
    "order_quantity": "下单数量",
    "trade_amount": "贸易额（美金）",
}

QUOTE_FIELD_LABELS = {
    "order_quantity": "订单数量",
    "calc_quantity": "算价格数量",
    "port_misc_fee_cny": "港杂费",
    "exchange_rate": "报价汇率",
    "commission_pct": "佣金",
    "customer_target_price_usd": "客户目标价",
    "current_exchange_rate": "当下汇率",
    "gross_profit_cny": "毛利润额",
    "final_quote_usd": "给客人报的价格",
    "selected_factory": "选用工厂",
    "selected_factory_price_cny": "选用工厂价格",
    "client_quoted_date": "给客人报价日期",
    "material_received_date": "收到资料日期",
    "price_tracking_notes": "价格追踪说明",
}


DOMESTIC_ROUNDS = {
    1: {
        "fields": {
            "order_quantity": "U",
            "calc_quantity": "V",
            "port_misc_fee_cny": "X",
            "exchange_rate": "AD",
            "commission_pct": "AF",
            "customer_target_price_usd": "BD",
            "current_exchange_rate": "BA",
            "gross_profit_cny": "BB",
            "final_quote_usd": "AZ",
            "selected_factory": "AX",
            "selected_factory_price_cny": "AY",
        },
        "factory_cols": ["AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT"],
    },
    2: {
        "fields": {
            "order_quantity": "BM",
            "calc_quantity": "BN",
            "port_misc_fee_cny": "BP",
            "exchange_rate": "BV",
            "commission_pct": "BX",
            "customer_target_price_usd": "CY",
            "current_exchange_rate": "CS",
            "gross_profit_cny": "CT",
            "final_quote_usd": "CR",
            "selected_factory": "CP",
            "selected_factory_price_cny": "CQ",
        },
        "factory_cols": ["CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL"],
    },
    3: {
        "fields": {
            "order_quantity": "DH",
            "calc_quantity": "DI",
            "port_misc_fee_cny": "DK",
            "exchange_rate": "DQ",
            "commission_pct": "DS",
            "customer_target_price_usd": "ET",
            "current_exchange_rate": "EN",
            "gross_profit_cny": "EO",
            "final_quote_usd": "EM",
            "selected_factory": "EK",
            "selected_factory_price_cny": "EL",
        },
        "factory_cols": ["EE", "EF", "EG"],
    },
    4: {
        "fields": {
            "order_quantity": "FC",
            "calc_quantity": "FD",
            "port_misc_fee_cny": "FF",
            "exchange_rate": "FL",
            "commission_pct": "FN",
            "customer_target_price_usd": "GO",
            "current_exchange_rate": "GI",
            "gross_profit_cny": "GJ",
            "final_quote_usd": "GH",
            "selected_factory": "GF",
            "selected_factory_price_cny": "GG",
        },
        "factory_cols": ["FZ", "GA", "GB"],
    },
}

TRACKING_ROUNDS = {
    1: {"material_received_date": "L", "client_quoted_date": "O", "price_tracking_notes": "Q", "domestic_arranged": "M", "overseas_arranged": "N"},
    2: {"material_received_date": "R", "client_quoted_date": "U", "price_tracking_notes": "W", "domestic_arranged": "S", "overseas_arranged": "T"},
    3: {"material_received_date": "X", "client_quoted_date": "AA", "price_tracking_notes": "AC", "domestic_arranged": "Y", "overseas_arranged": "Z"},
    4: {"material_received_date": "AD", "client_quoted_date": "AG", "price_tracking_notes": "AI", "domestic_arranged": "AE", "overseas_arranged": "AF"},
    5: {"material_received_date": "AJ", "client_quoted_date": "AM", "price_tracking_notes": "AO", "domestic_arranged": "AK", "overseas_arranged": "AL"},
    6: {"material_received_date": "AP", "client_quoted_date": "AS", "price_tracking_notes": "AU", "domestic_arranged": "AQ", "overseas_arranged": "AR"},
    7: {"material_received_date": "AV", "client_quoted_date": "AY", "price_tracking_notes": "BA", "domestic_arranged": "AW", "overseas_arranged": "AX"},
    8: {"material_received_date": "BB", "client_quoted_date": "BE", "price_tracking_notes": "BG", "domestic_arranged": "BC", "overseas_arranged": "BD"},
}


@dataclass
class ExcelValue:
    field: str
    label: str
    value: Any
    source_sheet: str
    source_cell: str


@dataclass
class FactoryQuoteCandidate:
    quote_type: str
    quote_round: int
    factory_name: str
    factory_price: Decimal
    currency: str
    price_unit: str
    source_sheet: str
    source_cell: str
    factory_id: uuid.UUID | None = None
    factory_matched: bool = False
    existing_id: uuid.UUID | None = None
    existing_price: Decimal | None = None
    status: str = "new"


@dataclass
class ParsedInquiry:
    inquiry_no: str | None
    excel_locations: list[str] = field(default_factory=list)
    inquiry_fields: dict[str, ExcelValue] = field(default_factory=dict)
    quote_fields: dict[tuple[str, int], dict[str, ExcelValue]] = field(default_factory=dict)
    factory_quotes: list[FactoryQuoteCandidate] = field(default_factory=list)
    needs_confirmation: list[dict[str, Any]] = field(default_factory=list)
    parse_errors: list[str] = field(default_factory=list)


def _clean_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in {"#N/A", "#DIV/0!", "#NUM!", "#VALUE!"}:
        return None
    return s


def _clean_optional(v: Any) -> str | None:
    s = _clean_str(v)
    if not s:
        return None
    if s in {"0", "0.0", "—", "-", "--", "——", "———", "————", "_", "__", "___", "/", "／"}:
        return None
    return s


def _clean_inquiry_no(v: Any) -> str | None:
    s = _clean_str(v)
    if not s:
        return None
    first = re.split(r"[\s（(]", s, maxsplit=1)[0].strip()
    if not re.search(r"[A-Za-z]", first):
        return None
    if not re.match(r"^[A-Za-z0-9][A-Za-z0-9_-]*$", first):
        return None
    return first


def _is_empty(v: Any) -> bool:
    return v is None or v == "" or str(v).strip() in {"", "#N/A", "#DIV/0!", "#NUM!", "#VALUE!"}


def _to_decimal(v: Any) -> Decimal | None:
    if _is_empty(v):
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_int(v: Any) -> int | None:
    if _is_empty(v):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v)
    total_match = re.search(r"共\s*(\d[\d,]*)\s*件", s)
    if total_match:
        return int(total_match.group(1).replace(",", ""))
    matches = [int(m.replace(",", "")) for m in re.findall(r"\d[\d,]*", s)]
    if not matches:
        return None
    return max(matches)


def _to_date(v: Any) -> date | None:
    if _is_empty(v):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def _clean_factory_name(v: Any) -> str | None:
    s = _clean_str(v)
    if not s:
        return None
    s = re.sub(r"[（(]\s*(元/件|/件)\s*[）)]", "", s)
    s = s.replace("工厂名字", "").strip()
    return s or None


def _cell(ws, row: int, col: str):
    return ws[f"{col}{row}"]


def _add_value(target: dict[str, ExcelValue], value: ExcelValue, parsed: ParsedInquiry) -> None:
    if _is_empty(value.value):
        return
    old = target.get(value.field)
    if old and str(old.value) != str(value.value):
        parsed.needs_confirmation.append({
            "field_name": value.label,
            "excel_value": value.value,
            "source_sheet": value.source_sheet,
            "source_cell": value.source_cell,
            "reason": f"同一询单同字段出现多个不同 Excel 值，已保留首次值 {old.value!r}",
            "suggestion": "后续人工确认是否按款式/场景拆分。",
        })
        return
    target[value.field] = value


def _parsed_for(records: dict[str, ParsedInquiry], inquiry_no: str | None) -> ParsedInquiry:
    key = inquiry_no or ""
    if key not in records:
        records[key] = ParsedInquiry(inquiry_no=inquiry_no)
    return records[key]


def _parse_workbook(file_bytes: bytes) -> tuple[dict[str, ParsedInquiry], dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    records: dict[str, ParsedInquiry] = {}
    sheet_stats: dict[str, Any] = {}

    if "总表" in wb.sheetnames:
        ws = wb["总表"]
        rows = 0
        for r in range(3, ws.max_row + 1):
            inquiry_no = _clean_inquiry_no(ws[f"A{r}"].value)
            if not inquiry_no:
                continue
            rows += 1
            parsed = _parsed_for(records, inquiry_no)
            parsed.excel_locations.append(f"总表!A{r}")

            for field, label, col, converter in (
                ("customer_order_no", "客户订单号", "B", _clean_optional),
                ("series_name", "系列", "C", _clean_optional),
                ("season", "季节", "D", _clean_optional),
                ("product_name", "品名", "F", _clean_optional),
                ("quantity", "数量", "L", _to_int),
                ("order_status", "订单状态", "T", _clean_optional),
                ("order_date", "下单日期", "GZ", _to_date),
                ("order_unit_price", "下单美金价格", "HM", _to_decimal),
                ("order_quantity", "下单数量", "HC", _to_int),
                ("trade_amount", "贸易额（美金）", "HU", _to_decimal),
            ):
                cell = _cell(ws, r, col)
                _add_value(parsed.inquiry_fields, ExcelValue(field, label, converter(cell.value), "总表", cell.coordinate), parsed)

            for field, label, col in (
                ("测试费", "测试费", "Y"),
                ("杂费", "杂费", "Z"),
                ("净利润值", "净利润值", "AE"),
                ("系列为 0 或 —— 的处理", "系列", "C"),
            ):
                val = _clean_str(_cell(ws, r, col).value)
                if val:
                    parsed.needs_confirmation.append({
                        "field_name": label,
                        "excel_value": val,
                        "source_sheet": "总表",
                        "source_cell": f"{col}{r}",
                        "reason": f"{field} 本阶段暂不导入",
                        "suggestion": "保留在预览中，后续确认语义后再映射。",
                    })

            for round_no, cfg in DOMESTIC_ROUNDS.items():
                qkey = (DOMESTIC, round_no)
                parsed.quote_fields.setdefault(qkey, {})
                for field, col in cfg["fields"].items():
                    label = QUOTE_FIELD_LABELS[field]
                    cell = _cell(ws, r, col)
                    converter = _to_int if field in {"order_quantity", "calc_quantity"} else (_clean_str if field == "selected_factory" else _to_decimal)
                    value = converter(cell.value)
                    if field == "selected_factory":
                        value = _clean_factory_name(value)
                    _add_value(parsed.quote_fields[qkey], ExcelValue(field, label, value, "总表", cell.coordinate), parsed)

                for col in cfg["factory_cols"]:
                    name = _clean_factory_name(ws[f"{col}2"].value)
                    price = _to_decimal(_cell(ws, r, col).value)
                    if name and price is not None:
                        parsed.factory_quotes.append(FactoryQuoteCandidate(
                            quote_type=DOMESTIC,
                            quote_round=round_no,
                            factory_name=name,
                            factory_price=price,
                            currency="CNY",
                            price_unit=PRICE_UNIT,
                            source_sheet="总表",
                            source_cell=f"{col}{r}",
                        ))
        sheet_stats["总表"] = {"rows": rows, "quote_rounds": sorted(DOMESTIC_ROUNDS)}

    if "询单追踪详情汇总表" in wb.sheetnames:
        ws = wb["询单追踪详情汇总表"]
        rows = 0
        for r in range(4, ws.max_row + 1):
            inquiry_no = _clean_inquiry_no(ws[f"A{r}"].value)
            if not inquiry_no:
                continue
            rows += 1
            parsed = _parsed_for(records, inquiry_no)
            parsed.excel_locations.append(f"询单追踪详情汇总表!A{r}")
            for round_no, cfg in TRACKING_ROUNDS.items():
                qkey = (DOMESTIC, round_no)
                parsed.quote_fields.setdefault(qkey, {})
                for field in ("material_received_date", "client_quoted_date", "price_tracking_notes"):
                    col = cfg[field]
                    cell = _cell(ws, r, col)
                    converter = _clean_str if field == "price_tracking_notes" else _to_date
                    _add_value(parsed.quote_fields[qkey], ExcelValue(field, QUOTE_FIELD_LABELS[field], converter(cell.value), "询单追踪详情汇总表", cell.coordinate), parsed)
                for pending_field, col in (("国内工厂安排报价日期", cfg["domestic_arranged"]), ("海外工厂安排报价日期", cfg["overseas_arranged"])):
                    val = _to_date(_cell(ws, r, col).value)
                    if val:
                        parsed.needs_confirmation.append({
                            "field_name": pending_field,
                            "excel_value": val.isoformat(),
                            "source_sheet": "询单追踪详情汇总表",
                            "source_cell": f"{col}{r}",
                            "reason": "当前 quote_items 只有一个 factory_arranged_date，无法区分国内/海外。",
                            "suggestion": "后续新增 domestic/overseas arranged date 后再导入。",
                        })
            for label, col in (("原样地点", "G"), ("原样样板参考情况", "I"), ("客供布色/品质样参考情况", "J")):
                val = _clean_str(_cell(ws, r, col).value)
                if val:
                    parsed.needs_confirmation.append({
                        "field_name": label,
                        "excel_value": val,
                        "source_sheet": "询单追踪详情汇总表",
                        "source_cell": f"{col}{r}",
                        "reason": "本阶段暂不导入原样/客供资料。",
                        "suggestion": "后续确认是否新增专字段或进入 remark。",
                    })
        sheet_stats["询单追踪详情汇总表"] = {"rows": rows, "quote_rounds": sorted(TRACKING_ROUNDS)}

    if "海外报价表-美金" in wb.sheetnames:
        ws = wb["海外报价表-美金"]
        rows = 0
        for r in range(5, ws.max_row + 1):
            inquiry_no = _clean_inquiry_no(ws[f"A{r}"].value)
            if not inquiry_no:
                continue
            rows += 1
            parsed = _parsed_for(records, inquiry_no)
            parsed.excel_locations.append(f"海外报价表-美金!A{r}")
            for field, label, col, converter in (
                ("customer_order_no", "客户订单号", "B", _clean_optional),
                ("product_name", "品名", "D", _clean_optional),
                ("quantity", "数量", "K", _to_int),
            ):
                cell = _cell(ws, r, col)
                _add_value(parsed.inquiry_fields, ExcelValue(field, label, converter(cell.value), "海外报价表-美金", cell.coordinate), parsed)
            qkey = (OVERSEAS, 1)
            parsed.quote_fields.setdefault(qkey, {})
            for field, label, col, converter in (
                ("calc_quantity", "算价格数量", "J", _to_int),
                ("order_quantity", "订单数量", "K", _to_int),
                ("commission_pct", "佣金", "O", _to_decimal),
                ("current_exchange_rate", "当下汇率", "AD", _to_decimal),
                ("customer_target_price_usd", "客户目标价", "AK", _to_decimal),
            ):
                cell = _cell(ws, r, col)
                _add_value(parsed.quote_fields[qkey], ExcelValue(field, label, converter(cell.value), "海外报价表-美金", cell.coordinate), parsed)
            val = _to_decimal(ws[f"AC{r}"].value)
            if val is not None:
                parsed.needs_confirmation.append({
                    "field_name": "海外给客人报的价格",
                    "excel_value": str(val),
                    "source_sheet": "海外报价表-美金",
                    "source_cell": f"AC{r}",
                    "reason": "是否覆盖国内对客报价尚未确认，本阶段不导入。",
                    "suggestion": "后续确认是否作为 overseas quote_items.final_quote_usd 单独保存。",
                })
            for col in ("P", "Q", "R"):
                name = _clean_factory_name(ws[f"{col}4"].value)
                price = _to_decimal(_cell(ws, r, col).value)
                if name and price is not None:
                    parsed.factory_quotes.append(FactoryQuoteCandidate(
                        quote_type=OVERSEAS,
                        quote_round=1,
                        factory_name=name,
                        factory_price=price,
                        currency="USD",
                        price_unit=PRICE_UNIT,
                        source_sheet="海外报价表-美金",
                        source_cell=f"{col}{r}",
                    ))
        sheet_stats["海外报价表-美金"] = {"rows": rows, "quote_rounds": [1]}

    return records, sheet_stats


def _json_value(v: Any) -> Any:
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, uuid.UUID):
        return str(v)
    return v


def _user_display_name(user: Any) -> str | None:
    return getattr(user, "display_name", None) or getattr(user, "username", None)


def _inquiry_fields_with_uploader(parsed: ParsedInquiry, user: Any) -> dict[str, ExcelValue]:
    fields = dict(parsed.inquiry_fields)
    uploader = _user_display_name(user)
    if uploader:
        fields.setdefault(
            "responsible_sales",
            ExcelValue("responsible_sales", "负责业务员", uploader, "当前账号", "当前登录用户"),
        )
    group_name = getattr(user, "group_name", None)
    if group_name:
        fields.setdefault(
            "group_name",
            ExcelValue("group_name", "所属小组", group_name, "当前账号", "当前登录用户"),
        )
    return fields


def _same_value(a: Any, b: Any) -> bool:
    if a is None and b in (None, ""):
        return True
    if isinstance(a, Decimal) or isinstance(b, Decimal):
        da = _to_decimal(a)
        db = _to_decimal(b)
        return da == db
    if isinstance(a, (date, datetime)) or isinstance(b, (date, datetime)):
        return _to_date(a) == _to_date(b)
    return str(a).strip() == str(b).strip()


async def _find_factory(db: AsyncSession, name: str) -> tuple[uuid.UUID | None, str]:
    row = (await db.execute(
        select(Factory).where(or_(Factory.factory_name == name, Factory.factory_short_name == name))
    )).scalars().first()
    if row:
        return row.id, row.factory_short_name or row.factory_name or name
    return None, name


async def _find_factory_quote(db: AsyncSession, inquiry_id: uuid.UUID, fq: FactoryQuoteCandidate) -> FactoryQuoteRecord | None:
    q = select(FactoryQuoteRecord).where(
        FactoryQuoteRecord.inquiry_id == inquiry_id,
        FactoryQuoteRecord.quote_round == fq.quote_round,
        func.coalesce(FactoryQuoteRecord.quote_type, DOMESTIC) == fq.quote_type,
    )
    if fq.factory_id:
        q = q.where(FactoryQuoteRecord.factory_id == fq.factory_id)
    else:
        q = q.where(
            FactoryQuoteRecord.factory_id.is_(None),
            func.lower(FactoryQuoteRecord.factory_name) == fq.factory_name.lower(),
        )
    return (await db.execute(q)).scalars().first()


async def _get_quote_item(db: AsyncSession, inquiry_id: uuid.UUID, quote_type: str, quote_round: int) -> QuoteItem | None:
    return (await db.execute(
        select(QuoteItem).where(
            QuoteItem.inquiry_id == inquiry_id,
            QuoteItem.quote_type == quote_type,
            QuoteItem.quote_round == quote_round,
        )
    )).scalars().first()


def _field_preview(field: ExcelValue, current: Any, key: str, system_table: str) -> dict[str, Any]:
    if _is_empty(field.value):
        status = "empty"
    elif current is None or current == "":
        status = "fillable"
    elif _same_value(current, field.value):
        status = "same"
    else:
        status = "conflict"
    return {
        "key": key,
        "field": field.field,
        "field_name": field.label,
        "system_table": system_table,
        "system_value": _json_value(current),
        "excel_value": _json_value(field.value),
        "source_sheet": field.source_sheet,
        "source_cell": field.source_cell,
        "status": status,
        "default_action": "keep_system" if status == "conflict" else "use_excel",
    }


async def preview_journey_import(db: AsyncSession, file_bytes: bytes, file_name: str, user: Any) -> dict[str, Any]:
    records, sheet_stats = _parse_workbook(file_bytes)
    rows: list[dict[str, Any]] = []
    totals = {
        "total_inquiries": 0, "matched": 0, "not_found": 0, "ambiguous": 0,
        "failed": 0, "ready_to_fill": 0, "conflict": 0,
        "factory_quote_conflicts": 0,
        "fillable_inquiry_fields": 0,
        "rows_with_fillable_fields": 0,
    }

    for inquiry_no, parsed in records.items():
        totals["total_inquiries"] += 1
        if not inquiry_no:
            totals["failed"] += 1
            rows.append({
                "inquiry_no": None, "inquiry_id": None, "status": "failed",
                "excel_locations": parsed.excel_locations, "errors": ["Excel 询单号为空"],
                "inquiry_fields": [], "quote_items": [], "factory_quotes": [],
                "needs_confirmation": parsed.needs_confirmation,
                "domestic_quote_rounds": 0, "overseas_quote_rounds": 0,
                "fillable_inquiry_fields": 0, "quote_items_to_create": 0,
                "factory_quotes_to_create": 0, "conflict_count": 0,
                "unmapped_count": len(parsed.needs_confirmation),
                "can_confirm": False,
            })
            continue

        matches = (await db.execute(select(Inquiry).where(Inquiry.inquiry_no == inquiry_no))).scalars().all()
        if not matches:
            totals["not_found"] += 1
            rows.append({
                "inquiry_no": inquiry_no, "inquiry_id": None, "status": "not_found",
                "excel_locations": parsed.excel_locations, "errors": ["系统未找到该询单号"],
                "inquiry_fields": [], "quote_items": [], "factory_quotes": [],
                "needs_confirmation": parsed.needs_confirmation,
                "domestic_quote_rounds": len([k for k in parsed.quote_fields if k[0] == DOMESTIC]),
                "overseas_quote_rounds": len([k for k in parsed.quote_fields if k[0] == OVERSEAS]),
                "fillable_inquiry_fields": 0, "quote_items_to_create": 0,
                "factory_quotes_to_create": len(parsed.factory_quotes), "conflict_count": 0,
                "unmapped_count": len(parsed.needs_confirmation), "can_confirm": False,
            })
            continue
        if len(matches) > 1:
            totals["ambiguous"] += 1
            rows.append({
                "inquiry_no": inquiry_no, "inquiry_id": None, "status": "ambiguous",
                "excel_locations": parsed.excel_locations, "errors": ["系统存在多个同编号询单"],
                "inquiry_fields": [], "quote_items": [], "factory_quotes": [],
                "needs_confirmation": parsed.needs_confirmation,
                "domestic_quote_rounds": 0, "overseas_quote_rounds": 0,
                "fillable_inquiry_fields": 0, "quote_items_to_create": 0,
                "factory_quotes_to_create": 0, "conflict_count": 0,
                "unmapped_count": len(parsed.needs_confirmation), "can_confirm": False,
            })
            continue

        inq = matches[0]
        if not can_edit_inquiry(inq, user):
            totals["failed"] += 1
            rows.append({
                "inquiry_no": inquiry_no, "inquiry_id": str(inq.id), "status": "failed",
                "excel_locations": parsed.excel_locations, "errors": ["无权更新该询单"],
                "inquiry_fields": [], "quote_items": [], "factory_quotes": [],
                "needs_confirmation": parsed.needs_confirmation,
                "domestic_quote_rounds": 0, "overseas_quote_rounds": 0,
                "fillable_inquiry_fields": 0, "quote_items_to_create": 0,
                "factory_quotes_to_create": 0, "conflict_count": 0,
                "unmapped_count": len(parsed.needs_confirmation), "can_confirm": False,
            })
            continue

        parsed_inquiry_fields = _inquiry_fields_with_uploader(parsed, user)
        inquiry_fields = [
            _field_preview(v, getattr(inq, k, None), f"{inquiry_no}|inquiry|{k}", "inquiries")
            for k, v in parsed_inquiry_fields.items()
        ]
        quote_items = []
        quote_items_to_create = 0
        for (quote_type, round_no), fields in sorted(parsed.quote_fields.items()):
            existing = await _get_quote_item(db, inq.id, quote_type, round_no)
            if existing is None and any(not _is_empty(v.value) for v in fields.values()):
                quote_items_to_create += 1
            quote_items.append({
                "quote_type": quote_type,
                "quote_round": round_no,
                "exists": existing is not None,
                "fields": [
                    _field_preview(v, getattr(existing, k, None) if existing else None, f"{inquiry_no}|quote|{quote_type}|{round_no}|{k}", "quote_items")
                    for k, v in fields.items()
                ],
            })

        factory_quotes = []
        for fq in parsed.factory_quotes:
            fq.factory_id, fq.factory_name = await _find_factory(db, fq.factory_name)
            fq.factory_matched = fq.factory_id is not None
            existing_fq = await _find_factory_quote(db, inq.id, fq)
            if existing_fq is None:
                status = "new"
            elif _same_value(existing_fq.factory_price, fq.factory_price):
                status = "same"
            else:
                status = "factory_quote_conflict"
                totals["factory_quote_conflicts"] += 1
            factory_quotes.append({
                "key": f"{inquiry_no}|factory|{fq.quote_type}|{fq.quote_round}|{fq.factory_id or fq.factory_name}",
                "quote_type": fq.quote_type,
                "quote_round": fq.quote_round,
                "factory_id": str(fq.factory_id) if fq.factory_id else None,
                "factory_name": fq.factory_name,
                "factory_matched": fq.factory_matched,
                "factory_price": float(fq.factory_price),
                "currency": fq.currency,
                "price_unit": fq.price_unit,
                "source_sheet": fq.source_sheet,
                "source_cell": fq.source_cell,
                "status": status,
                "system_price": float(existing_fq.factory_price) if existing_fq and existing_fq.factory_price is not None else None,
                "default_action": "keep_system" if status == "factory_quote_conflict" else "create",
                "message": None if fq.factory_matched else "未匹配工厂档案，将按名称保存报价记录",
            })

        conflict_count = sum(1 for f in inquiry_fields if f["status"] == "conflict")
        conflict_count += sum(1 for qi in quote_items for f in qi["fields"] if f["status"] == "conflict")
        conflict_count += sum(1 for f in factory_quotes if f["status"] == "factory_quote_conflict")
        fillable_inquiry = sum(1 for f in inquiry_fields if f["status"] == "fillable")
        factory_to_create = sum(1 for f in factory_quotes if f["status"] == "new")
        totals["fillable_inquiry_fields"] += fillable_inquiry
        if fillable_inquiry:
            totals["rows_with_fillable_fields"] += 1

        status = "conflict" if conflict_count else ("ready_to_fill" if fillable_inquiry or quote_items_to_create or factory_to_create else "matched")
        totals["matched"] += 1
        if status == "conflict":
            totals["conflict"] += 1
        if status == "ready_to_fill":
            totals["ready_to_fill"] += 1

        rows.append({
            "inquiry_no": inquiry_no,
            "inquiry_id": str(inq.id),
            "status": status,
            "excel_locations": parsed.excel_locations,
            "errors": [],
            "inquiry_fields": inquiry_fields,
            "quote_items": quote_items,
            "factory_quotes": factory_quotes,
            "needs_confirmation": parsed.needs_confirmation,
            "domestic_quote_rounds": len([k for k in parsed.quote_fields if k[0] == DOMESTIC]),
            "overseas_quote_rounds": len([k for k in parsed.quote_fields if k[0] == OVERSEAS]),
            "fillable_inquiry_fields": fillable_inquiry,
            "quote_items_to_create": quote_items_to_create,
            "factory_quotes_to_create": factory_to_create,
            "conflict_count": conflict_count,
            "unmapped_count": len(parsed.needs_confirmation),
            "can_confirm": True,
        })

    return {
        "file_name": file_name,
        "sheet_stats": sheet_stats,
        "summary": totals,
        "rows": rows,
    }


def _decision(decisions: dict[str, Any], key: str, default: str = "keep_system") -> str:
    return (decisions.get("fields", {}) or {}).get(key, default)


def _factory_decision(decisions: dict[str, Any], key: str, default: str = "keep_system") -> str:
    return (decisions.get("factory_quotes", {}) or {}).get(key, default)


async def _log_in_session(db: AsyncSession, user: Any, *, action_type: str, inquiry: Inquiry | None, target_type: str, target_id: Any = None, description: str, before: dict | None = None, after: dict | None = None, status: str = "success", error: str | None = None) -> None:
    db.add(OperationLog(
        id=uuid.uuid4(),
        **log_kwargs_from_user(user),
        action_type=action_type,
        target_type=target_type,
        target_id=str(target_id) if target_id is not None else None,
        inquiry_id=inquiry.id if inquiry else None,
        inquiry_no=inquiry.inquiry_no if inquiry else None,
        description=description,
        before_data_json=before,
        after_data_json=after,
        status=status,
        error_message=error,
    ))


async def confirm_journey_import(db: AsyncSession, file_bytes: bytes, file_name: str, user: Any, decisions: dict[str, Any] | str | None = None) -> dict[str, Any]:
    if isinstance(decisions, str):
        decisions = json.loads(decisions or "{}")
    decisions = decisions or {}
    records, sheet_stats = _parse_workbook(file_bytes)

    summary = {
        "updated_inquiries": 0, "filled_fields": 0, "updated_fields": 0,
        "created_quote_items": 0, "updated_quote_items": 0,
        "created_factory_quotes": 0, "updated_factory_quotes": 0,
        "skipped_conflicts": 0, "same_factory_quotes": 0,
        "not_found": 0, "ambiguous": 0, "failed": 0,
        "row_errors": [],
    }

    for inquiry_no, parsed in records.items():
        if not inquiry_no:
            summary["failed"] += 1
            summary["row_errors"].append({"inquiry_no": None, "error": "Excel 询单号为空"})
            continue

        matches = (await db.execute(select(Inquiry).where(Inquiry.inquiry_no == inquiry_no))).scalars().all()
        if not matches:
            summary["not_found"] += 1
            continue
        if len(matches) > 1:
            summary["ambiguous"] += 1
            continue
        inq = matches[0]
        if not can_edit_inquiry(inq, user):
            summary["failed"] += 1
            summary["row_errors"].append({"inquiry_no": inquiry_no, "error": "无权更新该询单"})
            continue

        try:
            async with db.begin_nested():
                inquiry_updates: dict[str, Any] = {}
                for field, excel in _inquiry_fields_with_uploader(parsed, user).items():
                    if _is_empty(excel.value):
                        continue
                    cur = getattr(inq, field, None)
                    key = f"{inquiry_no}|inquiry|{field}"
                    if cur is None or cur == "":
                        inquiry_updates[field] = excel.value
                        summary["filled_fields"] += 1
                    elif _same_value(cur, excel.value):
                        continue
                    elif _decision(decisions, key) == "excel":
                        inquiry_updates[field] = excel.value
                        summary["updated_fields"] += 1
                    else:
                        summary["skipped_conflicts"] += 1
                        await _log_in_session(db, user, action_type="inquiry_journey_import_field_conflict_skip", inquiry=inq, target_type="inquiry", target_id=inq.id, description=f"跳过冲突字段：{INQUIRY_FIELD_LABELS.get(field, field)}", before={field: _json_value(cur)}, after={field: _json_value(excel.value)})
                if inquiry_updates:
                    before = {k: _json_value(getattr(inq, k, None)) for k in inquiry_updates}
                    await crud.update_inquiry(db, inq.id, inquiry_updates)
                    for k, v in inquiry_updates.items():
                        setattr(inq, k, v)
                    summary["updated_inquiries"] += 1
                    await _log_in_session(db, user, action_type="inquiry_journey_import_field_update", inquiry=inq, target_type="inquiry", target_id=inq.id, description="来龙去脉 Excel 回填询单字段", before=before, after={k: _json_value(v) for k, v in inquiry_updates.items()})

                for (quote_type, round_no), fields in parsed.quote_fields.items():
                    if not any(not _is_empty(v.value) for v in fields.values()):
                        continue
                    item = await _get_quote_item(db, inq.id, quote_type, round_no)
                    created = False
                    if item is None:
                        item = QuoteItem(id=uuid.uuid4(), inquiry_id=inq.id, quote_type=quote_type, quote_round=round_no)
                        db.add(item)
                        await db.flush()
                        created = True
                        summary["created_quote_items"] += 1
                    updates: dict[str, Any] = {}
                    for field, excel in fields.items():
                        if _is_empty(excel.value):
                            continue
                        cur = getattr(item, field, None)
                        key = f"{inquiry_no}|quote|{quote_type}|{round_no}|{field}"
                        if cur is None or cur == "":
                            updates[field] = excel.value
                        elif _same_value(cur, excel.value):
                            continue
                        elif _decision(decisions, key) == "excel":
                            updates[field] = excel.value
                        else:
                            summary["skipped_conflicts"] += 1
                    if updates:
                        before = {k: _json_value(getattr(item, k, None)) for k in updates}
                        for k, v in updates.items():
                            setattr(item, k, v)
                        if not created:
                            summary["updated_quote_items"] += 1
                        await _log_in_session(db, user, action_type="inquiry_journey_import_field_update", inquiry=inq, target_type="quote_item", target_id=item.id, description=f"来龙去脉 Excel 回填 {quote_type} 第{round_no}轮报价字段", before=before, after={k: _json_value(v) for k, v in updates.items()})

                for fq in parsed.factory_quotes:
                    fq.factory_id, fq.factory_name = await _find_factory(db, fq.factory_name)
                    existing = await _find_factory_quote(db, inq.id, fq)
                    key = f"{inquiry_no}|factory|{fq.quote_type}|{fq.quote_round}|{fq.factory_id or fq.factory_name}"
                    if existing is None:
                        rec = FactoryQuoteRecord(
                            id=uuid.uuid4(),
                            inquiry_id=inq.id,
                            inquiry_no=inq.inquiry_no,
                            factory_id=fq.factory_id,
                            factory_name=fq.factory_name,
                            quote_type=fq.quote_type,
                            quote_round=fq.quote_round,
                            factory_price=fq.factory_price,
                            currency=fq.currency,
                            price_unit=fq.price_unit,
                            source_sheet=fq.source_sheet,
                            source_cell=fq.source_cell,
                            quoted_by=user.username,
                            quoted_at=datetime.now(timezone.utc),
                            created_by=user.username,
                        )
                        db.add(rec)
                        await db.flush()
                        summary["created_factory_quotes"] += 1
                        await _log_in_session(db, user, action_type="factory_quote_import_create", inquiry=inq, target_type="factory_quote", target_id=rec.id, description=f"导入工厂报价：{fq.quote_type} 第{fq.quote_round}轮 {fq.factory_name}", after={"factory_name": fq.factory_name, "factory_price": float(fq.factory_price), "currency": fq.currency, "source_sheet": fq.source_sheet, "source_cell": fq.source_cell})
                    elif _same_value(existing.factory_price, fq.factory_price):
                        summary["same_factory_quotes"] += 1
                    elif _factory_decision(decisions, key) == "use_excel":
                        before = {"factory_price": _json_value(existing.factory_price)}
                        existing.factory_price = fq.factory_price
                        existing.source_sheet = fq.source_sheet
                        existing.source_cell = fq.source_cell
                        existing.quoted_by = user.username
                        existing.quoted_at = datetime.now(timezone.utc)
                        summary["updated_factory_quotes"] += 1
                        await _log_in_session(db, user, action_type="factory_quote_import_update", inquiry=inq, target_type="factory_quote", target_id=existing.id, description=f"使用 Excel 价格更新工厂报价：{fq.quote_type} 第{fq.quote_round}轮 {fq.factory_name}", before=before, after={"factory_price": float(fq.factory_price), "source_sheet": fq.source_sheet, "source_cell": fq.source_cell})
                    elif _factory_decision(decisions, key) == "add_remark":
                        before = {"remark": existing.remark}
                        note = f"Excel导入冲突价格：{fq.factory_price} {fq.currency}/{fq.price_unit}（{fq.source_sheet}!{fq.source_cell}），系统价格已保留。"
                        existing.remark = f"{existing.remark}\n{note}" if existing.remark else note
                        existing.source_sheet = fq.source_sheet
                        existing.source_cell = fq.source_cell
                        summary["updated_factory_quotes"] += 1
                        await _log_in_session(db, user, action_type="factory_quote_import_update", inquiry=inq, target_type="factory_quote", target_id=existing.id, description=f"将 Excel 冲突工厂报价追加到备注：{fq.quote_type} 第{fq.quote_round}轮 {fq.factory_name}", before=before, after={"remark": existing.remark})
                    else:
                        summary["skipped_conflicts"] += 1
                        await _log_in_session(db, user, action_type="factory_quote_import_conflict_skip", inquiry=inq, target_type="factory_quote", target_id=existing.id, description=f"保留系统工厂报价：{fq.quote_type} 第{fq.quote_round}轮 {fq.factory_name}", before={"factory_price": _json_value(existing.factory_price)}, after={"excel_factory_price": float(fq.factory_price), "choice": "keep_system"})
        except Exception as exc:
            summary["failed"] += 1
            summary["row_errors"].append({"inquiry_no": inquiry_no, "error": str(exc), "locations": parsed.excel_locations})

    await safe_log(
        **log_kwargs_from_user(user),
        action_type="inquiry_journey_import_confirm",
        target_type="inquiry_journey_import",
        description="确认导入来龙去脉 Excel",
        after_data={"file_name": file_name, **summary},
    )
    return {"file_name": file_name, "sheet_stats": sheet_stats, "summary": summary}
