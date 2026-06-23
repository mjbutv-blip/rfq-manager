"""
询单表 Excel 解析器（纯解析层，不访问数据库）

主要入口：
  parse_excel_file(file_bytes, file_name) -> ParseResult

解析流程：
  1. 加载 workbook（data_only=True，读取计算结果而非公式）
  2. 选择目标 sheet（优先含"询单"/"总表"/"全公司"的 sheet）
  3. 自动识别表头行（扫描前 HEADER_SCAN_ROWS 行，优先找含"询单号"的行）
  4. 构建列映射（列索引 → 字段名 / 中文表头）
  5. 逐行解析：类型转换 + 必填校验 + 派生 inquiry_year/inquiry_month
  6. 返回 ParseResult（每行 status = valid | failed）

注意：本模块不做"同一询单号/同一款式重复"判断——那需要知道数据库里
已有什么、以及同批次内前面的行是否真正写库成功，纯解析层拿不到这些信息。
重复判断统一在 services/import_service.py 的 _classify_existing /
_write_valid_row 中按"实际写库结果"实时进行，避免"分类即占用"导致
前一行写库失败时，后面真正应该成功的同款行被错误地当作 duplicate_item。
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Literal

import openpyxl
import xlrd

from app.core.category_detect import detect_product_category
from app.core.field_mapping import (
    CUSTOMER_IDENTITY_FIELDS,
    DATE_FIELDS,
    DECIMAL_FIELDS,
    FIELD_MAPPING,
    FORMAL_REQUIRED_FIELDS,
    FORMAL_SKIP,
    FORMAL_TEMPLATE_FIELD_MAPPING,
    INT_FIELDS,
    ITEM_FIELD_MAPPING,
    PCT_FIELDS,
    REQUIRED_FIELDS,
)

HEADER_SCAN_ROWS = 15

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# ── 数据类 ─────────────────────────────────────────────────────────────────────

@dataclass
class ParsedRow:
    """单行解析结果（DB 状态未知，只经过本地校验）"""

    row_number: int
    inquiry_no: str | None
    # valid  = 本地校验通过（仍需 DB 层判断 new/existing/重复，见模块顶部说明）
    # failed = 必填字段缺失或类型转换失败
    status: Literal["valid", "failed"]
    raw_data: dict[str, Any]     # {中文表头: 原始值}
    parsed_data: dict[str, Any]  # {字段名: 转换后值}
    errors: list[str]

    @property
    def is_valid(self) -> bool:
        return self.status == "valid"

    @property
    def error_message(self) -> str | None:
        return "; ".join(self.errors) if self.errors else None


@dataclass
class ParseResult:
    """整个 Excel 文件的解析结果（不含 DB new/existing 状态）"""

    file_name: str
    sheet_name: str
    rows: list[ParsedRow]
    column_mapping: dict[str, str]   # {字段名: Excel 列头原始文字}，供前端展示
    missing_headers: list[str]       # REQUIRED_FIELDS 中在 Excel 没有对应列的字段名
    unmapped_headers: list[str]      # Excel 列头未匹配到 FIELD_MAPPING 的原始文字

    # ── 统计属性 ──────────────────────────────────────────────────────────────

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    @property
    def valid_count(self) -> int:
        return sum(1 for r in self.rows if r.status == "valid")

    @property
    def failed_count(self) -> int:
        return sum(1 for r in self.rows if r.status == "failed")

    def get_rows_by_status(self, status: str) -> list[ParsedRow]:
        return [r for r in self.rows if r.status == status]


# ── 日期解析 ────────────────────────────────────────────────────────────────────

def _parse_date(val: Any) -> tuple[date | None, str | None]:
    """返回 (date, error_message)。两者最多有一个非 None。"""
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.date(), None
    if isinstance(val, date):
        return val, None

    s = str(val).strip()
    if not s:
        return None, None

    for fmt in (
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
        "%d/%m/%Y", "%m/%d/%Y",
        "%Y年%m月%d日", "%Y%m%d",
    ):
        try:
            return datetime.strptime(s, fmt).date(), None
        except ValueError:
            continue

    # Excel 日期序列号（整数形式存为字符串的情况）
    try:
        n = int(float(s))
        if 30000 < n < 70000:
            from datetime import timedelta
            return date(1899, 12, 30) + timedelta(days=n), None
    except (ValueError, OverflowError):
        pass

    return None, f"日期格式无法识别：{val!r}（支持 YYYY-MM-DD / YYYY/MM/DD 等）"


# ── 数值解析 ────────────────────────────────────────────────────────────────────

def _parse_decimal(val: Any) -> tuple[Decimal | None, str | None]:
    """金额/数值 → Decimal，返回 (value, error_message)。"""
    if val is None:
        return None, None
    if isinstance(val, Decimal):
        return val, None
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val)), None
        except InvalidOperation:
            return None, f"数值无法解析：{val!r}"

    s = str(val).strip()
    if not s:
        return None, None
    cleaned = re.sub(r"[^\d.\-]", "", s)
    if not cleaned:
        return None, None
    try:
        return Decimal(cleaned), None
    except InvalidOperation:
        return None, f"数值无法解析：{val!r}"


def _parse_pct(val: Any) -> tuple[Decimal | None, str | None]:
    """
    百分比字段（毛利率）。
    支持：18.5 / 18.5% / 18.50%
    去除 % 后按原始数字保存，不做 0.x → x*100 自动换算。
    """
    if val is None:
        return None, None
    s = str(val).strip()
    if not s:
        return None, None
    cleaned = re.sub(r"[^\d.\-]", "", s)
    if not cleaned:
        return None, None
    try:
        return Decimal(cleaned), None
    except InvalidOperation:
        return None, f"百分比无法解析：{val!r}（支持 18.5、18.5%、18.50%）"


def _parse_int(val: Any) -> tuple[int | None, str | None]:
    """整数字段（数量等）。"""
    if val is None:
        return None, None
    if isinstance(val, int):
        return val, None
    if isinstance(val, float):
        return int(val), None

    s = str(val).strip()
    if not s:
        return None, None
    cleaned = re.sub(r"[^\d\-]", "", s)
    if not cleaned:
        return None, None
    try:
        return int(cleaned), None
    except ValueError:
        return None, f"整数无法解析：{val!r}"


def _coerce_field(field_name: str, raw_val: Any) -> tuple[Any, str | None]:
    """根据字段类型做类型转换，返回 (转换后值, 错误信息)。"""
    if raw_val is None or str(raw_val).strip() == "":
        return None, None

    if field_name in DATE_FIELDS:
        return _parse_date(raw_val)

    if field_name in INT_FIELDS:
        v, err = _parse_int(raw_val)
        if err:
            err = f"{field_name}: {err}"
        return v, err

    if field_name in DECIMAL_FIELDS:
        v, err = _parse_decimal(raw_val)
        if err:
            err = f"{field_name}: {err}"
        return v, err

    if field_name in PCT_FIELDS:
        v, err = _parse_pct(raw_val)
        if err:
            err = f"{field_name}: {err}"
        return v, err

    # 字符串字段：strip，截断超长内容
    return str(raw_val).strip()[:2000], None


# ── 正式报价单模板专用解析器 ─────────────────────────────────────────────────────

def _parse_text_quantity(val: Any) -> tuple[int | None, str | None]:
    """
    解析中文描述性数量文本，如"00单：共34849件，分两批走货"。
    取最后一个"共X件/套/卡"数字；无则取最后一个不跟"1卡"的"X件/套"数字。
    """
    if val is None:
        return None, None
    if isinstance(val, int):
        return val, None
    if isinstance(val, float):
        return int(val), None
    s = str(val).strip()
    if not s:
        return None, None
    # 优先匹配"共X件/套"（合计数），避免误取分批小计或"共X卡"
    matches = re.findall(r'共(\d+)[件套]', s)
    if matches:
        return int(matches[-1]), None
    # 其次取最后一个"X件/套"且后面不是"1卡"（排除"2件1卡"中的件数）
    matches = re.findall(r'(\d+)[件套](?!1卡)', s)
    if matches:
        return int(matches[-1]), None
    # 退而求其次取"共X卡"
    matches = re.findall(r'共(\d+)卡', s)
    if matches:
        return int(matches[-1]), None
    # 兜底：只有当文本里最多含一段数字时才当作普通数值处理；
    # 含多段数字的描述性文本（如多批次拆分说明）不做整数拼接，
    # 避免产生超出 int32 范围的乱码数字，改为标记解析失败
    numbers = re.findall(r'\d+', s)
    if len(numbers) <= 1:
        return _parse_int(val)
    return None, f"数量文本无法识别件数/套数：{val!r}"


def _parse_price_text(val: Any) -> tuple[Decimal | None, str | None]:
    """
    解析"13元/件"或"12.8元/卡"格式的价格文字，提取第一个数字。
    占位符（"—"/"——"等）返回 None。
    """
    if val is None:
        return None, None
    if isinstance(val, (int, float, Decimal)):
        return _parse_decimal(val)
    s = str(val).strip()
    if not s:
        return None, None
    match = re.search(r'(\d+(?:\.\d+)?)\s*元', s)
    if match:
        try:
            return Decimal(match.group(1)), None
        except InvalidOperation:
            pass
    if re.fullmatch(r'[-—\s]+', s):
        return None, None
    return _parse_decimal(val)


def _is_formal_quotation_template(ws) -> bool:
    """
    检测是否为正式报价单汇总模板：
    - 当前 sheet 名称为"总表"
    - 前若干行中存在含"FOB"的单元格（即 FOB 价格列头）
    """
    if ws.title != "总表":
        return False
    for row_idx in range(1, min(HEADER_SCAN_ROWS, ws.max_row) + 1):
        for cell in ws[row_idx]:
            if cell.value and "FOB" in str(cell.value):
                return True
    return False


def _extract_country_from_sheet(ws) -> str | None:
    """
    从总表 sheet 标题行提取客户国家。
    标题格式：'TK-BTKU1005-1013报价单    德国' → '德国'
    """
    for row_idx in range(1, 5):
        for cell in ws[row_idx]:
            if cell.value and '报价单' in str(cell.value):
                # 取"报价单"后面的第一个词（国家名）
                m = re.search(r'报价单\s+(\S+)', str(cell.value))
                if m:
                    candidate = m.group(1).strip()
                    # 必须含中文（排除括号、数字等噪音）
                    if re.search(r'[一-鿿]', candidate):
                        # 只取中文部分（去掉末尾可能跟着的括号等）
                        return re.match(r'[一-鿿]+', candidate).group()
    return None


def _fill_formal_defaults(
    row: "ParsedRow",
    scope_user: Any,
    country: str | None,
) -> None:
    """
    正式模板行：将用户上下文和 sheet 标题推导出的字段写入 parsed_data。
    （responsible_sales 由上传账号强制覆盖，见 _parse_single_row 的
    force_responsible_sales 参数，此处不再处理。）
    - sheet 标题内国家词 → country
    - 登录用户组别       → group_name
    """
    if scope_user is None:
        return
    if "group_name" not in row.parsed_data and getattr(scope_user, "group_name", None):
        row.parsed_data["group_name"] = scope_user.group_name
    if country and "country" not in row.parsed_data:
        row.parsed_data["country"] = country


# ── 老版 .xls（二进制 OLE2）兼容 ─────────────────────────────────────────────────

def _xls_cell_value(cell: "xlrd.sheet.Cell", book: "xlrd.book.Book") -> Any:
    """把 xlrd 单元格值转换为与 openpyxl 一致的 Python 类型。"""
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        # 整数值用 int 表示，避免 18.0 这种多余的小数点
        return int(cell.value) if cell.value == int(cell.value) else cell.value
    return cell.value


def _materialize_readonly_workbook(ro_wb: openpyxl.Workbook) -> openpyxl.Workbook:
    """
    把 read_only 模式加载的 Workbook 转存为普通（可随机访问）的 openpyxl
    Workbook。read_only 模式下的单元格（含 EmptyCell 占位）没有 .row/.column
    属性，与后续解析逻辑（依赖 cell.column）不兼容，需要先转换。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for ro_ws in ro_wb.worksheets:
        ws = wb.create_sheet(title=ro_ws.title)
        for r, row_cells in enumerate(ro_ws.iter_rows(), start=1):
            for c, cell in enumerate(row_cells, start=1):
                if cell.value is not None:
                    ws.cell(row=r, column=c, value=cell.value)
    return wb


def _load_workbook(file_bytes: bytes) -> openpyxl.Workbook:
    """
    加载 Excel 文件为 openpyxl.Workbook。
    自动识别老版二进制 .xls（OLE2 格式）并通过 xlrd 转换为等价的
    openpyxl Workbook，使后续解析逻辑（依赖 openpyxl 的 cell/iter_rows API）
    无需区分文件格式。
    """
    if file_bytes[:4] == b"PK\x03\x04":
        try:
            return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except ValueError:
            # 部分由 WPS 等工具生成的文件，其内嵌图片/图表的 XML 不规范
            # （如 panose 字体属性格式错误），导致 openpyxl 完整加载失败。
            # 读取数据并不需要图片/图表，改用 read_only 模式可绕过该部分解析，
            # 再转存为普通 Workbook 以保持与后续解析逻辑的兼容。
            ro_wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            return _materialize_readonly_workbook(ro_wb)

    if file_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        xls_book = xlrd.open_workbook(file_contents=file_bytes)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet in xls_book.sheets():
            ws = wb.create_sheet(title=sheet.name)
            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    value = _xls_cell_value(cell, xls_book)
                    if value is not None:
                        ws.cell(row=r + 1, column=c + 1, value=value)
        return wb

    raise ValueError("文件不是有效的 Excel 文件（既不是 .xlsx 也不是 .xls 格式）")


# ── Sheet / 表头识别 ────────────────────────────────────────────────────────────

def _select_sheet(wb: openpyxl.Workbook) -> str:
    """优先选择名称含"询单"/"总表"/"全公司"的 sheet，否则取第一个。"""
    preferred = [
        name for name in wb.sheetnames
        if any(kw in name for kw in ("询单", "总表", "全公司"))
    ]
    return preferred[0] if preferred else wb.sheetnames[0]


def _find_header_row(ws, mapping: dict[str, str] | None = None) -> int:
    """
    扫描前 HEADER_SCAN_ROWS 行：
    - 优先返回含"询单号"的行
    - 次选命中 mapping 关键词最多的行
    - 找不到则返回第 1 行
    """
    known_keys = set((mapping or FIELD_MAPPING).keys())
    best_row, best_count = 1, 0

    for row_idx in range(1, HEADER_SCAN_ROWS + 1):
        cell_texts = [
            str(c.value or "").strip().replace("\n", "")
            for c in ws[row_idx]
        ]

        if any("询单号" in t for t in cell_texts):
            return row_idx

        count = sum(
            1 for t in cell_texts
            if any(t.startswith(k) or k in t for k in known_keys)
        )
        if count > best_count:
            best_count, best_row = count, row_idx

    return best_row


def _build_col_indices(
    ws, header_row: int, mapping: dict[str, str] | None = None
) -> tuple[dict[int, str], dict[int, str]]:
    """
    遍历表头行，返回：
    - col_field : {列索引: 字段名}
    - col_header: {列索引: 中文表头原始文字}
    """
    if mapping is None:
        mapping = FIELD_MAPPING

    col_field: dict[int, str] = {}
    col_header: dict[int, str] = {}

    # 按关键词长度降序排列，让较长的关键词（如"下单数量"）
    # 优先于较短的（如"数量"），避免误匹配
    sorted_mapping = sorted(mapping.items(), key=lambda x: len(x[0]), reverse=True)

    for cell in ws[header_row]:
        raw = str(cell.value or "").strip().replace("\n", "")
        if not raw:
            continue
        col_header[cell.column] = raw

        for keyword, field_name in sorted_mapping:
            if raw.startswith(keyword) or keyword in raw:
                # FORMAL_SKIP 明确标记为跳过的列不写入 col_field
                if cell.column not in col_field and field_name != FORMAL_SKIP:
                    col_field[cell.column] = field_name
                break

    return col_field, col_header


# ── 单行解析 ───────────────────────────────────────────────────────────────────

def _derive_year_month(d: date | None) -> tuple[int | None, str | None]:
    if not d:
        return None, None
    return d.year, MONTH_ABBR[d.month - 1]


def build_item_identity_key(row_data: dict[str, Any]) -> tuple | None:
    """
    款式身份键：用于判断"同一询单号下是否出现了相同款式"。

    优先级：
      1. inquiry_no + style_no（两者都非空）
      2. inquiry_no + product_name + series_name（三者都非空）
      3. 无法构建可靠键时返回 None —— 调用方不应把 None 当作"重复"或
         "新款"，而应单独标记为"无法判断"，等待人工确认。

    纯函数，不访问数据库；import_service 在此基础上叠加 DB 层面的
    新旧询单/新旧款式判断。
    """
    inquiry_no = str(row_data.get("inquiry_no") or "").strip()
    if not inquiry_no:
        return None

    style_no = str(row_data.get("style_no") or "").strip()
    if style_no:
        return (inquiry_no, "style_no", style_no)

    product_name = str(row_data.get("product_name") or "").strip()
    series_name = str(row_data.get("series_name") or "").strip()
    if product_name and series_name:
        return (inquiry_no, "name", product_name, series_name)

    return None


def _parse_single_row(
    row_number: int,
    row_cells: tuple,
    col_field: dict[int, str],
    col_header: dict[int, str],
    required_fields: list[str] | None = None,
    customer_identity_fields: list[str] | None = None,
    extra_coerce: dict[str, Any] | None = None,
    force_responsible_sales: str | None = None,
) -> ParsedRow:
    """
    解析单行；status 暂定为 valid 或 failed（duplicate 在后续步骤检测）。
    extra_coerce: {field_name: callable(val) -> (value, error)} 覆盖特定字段的类型转换。
    force_responsible_sales: 非空时强制覆盖 responsible_sales（哪个账号上传就写谁），
    忽略 Excel 里原有的值。
    """
    if required_fields is None:
        required_fields = REQUIRED_FIELDS
    if customer_identity_fields is None:
        customer_identity_fields = CUSTOMER_IDENTITY_FIELDS

    raw_data: dict[str, Any] = {}
    parsed_data: dict[str, Any] = {}
    errors: list[str] = []

    for cell in row_cells:
        col = cell.column
        raw_header = col_header.get(col)
        field_name = col_field.get(col)

        if not raw_header:
            continue

        # raw_data 保存所有有值的原始中文列头 → 值（含未映射的列）
        if cell.value is not None:
            raw_data[raw_header] = cell.value

        if not field_name:
            continue

        if extra_coerce and field_name in extra_coerce:
            val, err = extra_coerce[field_name](cell.value)
        else:
            val, err = _coerce_field(field_name, cell.value)

        if err:
            errors.append(err)
        elif val is not None:
            parsed_data[field_name] = val

    # 哪个账号上传就写谁：强制覆盖 Excel 里的负责业务员，须在必填校验前完成
    if force_responsible_sales:
        parsed_data["responsible_sales"] = force_responsible_sales

    # 必填字段校验（用 is None 判断，允许数值 0 通过）
    for req in required_fields:
        val = parsed_data.get(req)
        missing = val is None or (isinstance(val, str) and not val.strip())
        if missing:
            errors.append(f"缺少必填字段：{req}")

    # 客户标识：至少一个不为空（字符串字段，空字符串等同于缺失）
    if customer_identity_fields and not any(parsed_data.get(f) for f in customer_identity_fields):
        errors.append(
            "客户标识缺失：%s 至少填写一个" % " 或 ".join(customer_identity_fields)
        )

    # 派生 inquiry_year / inquiry_month（不算必填，只是补充）
    inq_date = parsed_data.get("inquiry_date")
    if isinstance(inq_date, date):
        year, month = _derive_year_month(inq_date)
        if year:
            parsed_data["inquiry_year"] = year
        if month:
            parsed_data["inquiry_month"] = month

    status: Literal["valid", "failed"] = "failed" if errors else "valid"
    inquiry_no = str(parsed_data.get("inquiry_no", "")).strip() or None

    return ParsedRow(
        row_number=row_number,
        inquiry_no=inquiry_no,
        status=status,
        raw_data=raw_data,
        parsed_data=parsed_data,
        errors=errors,
    )


# ── 主入口 ─────────────────────────────────────────────────────────────────────

def parse_excel_file(
    file_bytes: bytes,
    file_name: str,
    scope_user: Any = None,
    override_sales: str | None = None,
) -> ParseResult:
    """
    解析询单表 Excel，不访问数据库。
    自动检测正式报价单汇总模板（总表 sheet + FOB 价格列），
    检测到时使用宽松必填规则并从 scope_user 补充组别/业务员默认值。
    override_sales 非空时优先级最高，强制覆盖全部行的 responsible_sales
    （调用方负责校验权限，此处不做角色检查）。

    返回 ParseResult，每行 status：
      valid  — 通过本地校验，是否新询单/已有询单/重复款式由 import_service
               在写库阶段按实际数据库状态实时判断（见模块顶部说明）
      failed — 必填字段缺失或类型转换失败

    Raises:
      ValueError  — 无法识别任何列头
      openpyxl 抛出的异常 — 文件损坏或格式不支持
    """
    wb = _load_workbook(file_bytes)
    sheet_name = _select_sheet(wb)
    ws = wb[sheet_name]

    # ── 检测模板类型 ────────────────────────────────────────────────────────────
    is_formal = _is_formal_quotation_template(ws)
    # 询单级映射叠加款式级映射（style_no / 工艺 / 尺码等），同一行 Excel
    # 既要解析询单级字段也要解析款式级字段，写库时再由 import_service 拆分。
    active_mapping = {
        **(FORMAL_TEMPLATE_FIELD_MAPPING if is_formal else FIELD_MAPPING),
        **ITEM_FIELD_MAPPING,
    }
    active_required = FORMAL_REQUIRED_FIELDS if is_formal else list(REQUIRED_FIELDS)
    active_identity = [] if is_formal else list(CUSTOMER_IDENTITY_FIELDS)
    row_extra_coerce: dict[str, Any] | None = (
        {"quantity": _parse_text_quantity, "final_quote": _parse_price_text}
        if is_formal else None
    )

    header_row = _find_header_row(ws, active_mapping)
    col_field, col_header = _build_col_indices(ws, header_row, active_mapping)

    if not col_field:
        raise ValueError(
            f"未能识别任何列头，请确认 Excel 列头文字与系统字段映射一致。"
            f"（文件：{file_name}，Sheet：{ws.title}，扫描了前 {HEADER_SCAN_ROWS} 行）"
        )

    # 字段名 → Excel 列头（供前端展示列映射关系）
    column_mapping: dict[str, str] = {}
    for col_idx, fn in col_field.items():
        if fn not in column_mapping:
            column_mapping[fn] = col_header.get(col_idx, fn)

    # 缺失的必填表头
    all_required = active_required + active_identity
    missing_headers = [f for f in all_required if f not in column_mapping]

    # 未映射的 Excel 列头
    mapped_raw_headers = {col_header[c] for c in col_field}
    all_raw_headers = set(col_header.values())
    unmapped_headers = sorted(all_raw_headers - mapped_raw_headers)

    # 业务员账号自己上传：强制写为本人，防止弄错/冒充他人。
    # 管理员/组长上传：通常是帮业务员代传，不强制覆盖，按 Excel 里写的人名为准，
    # 只有 Excel 没写时才用上传账号本人补默认值。
    uploader_name = (
        getattr(scope_user, "display_name", None) or getattr(scope_user, "username", None)
        if scope_user is not None else None
    )
    is_sales_uploader = scope_user is not None and getattr(scope_user, "role", None) == "sales"
    force_responsible_sales = override_sales or (uploader_name if is_sales_uploader else None)
    default_responsible_sales = None if (override_sales or is_sales_uploader) else uploader_name

    # 逐行解析
    rows: list[ParsedRow] = []
    for row_cells in ws.iter_rows(min_row=header_row + 1):
        row_num = row_cells[0].row

        # 全空行跳过
        if all(c.value is None for c in row_cells):
            continue

        # 跳过重复表头行
        first_val = str(row_cells[0].value or "").strip()
        if first_val in active_mapping:
            continue

        pr = _parse_single_row(
            row_num, row_cells, col_field, col_header,
            required_fields=active_required,
            customer_identity_fields=active_identity,
            extra_coerce=row_extra_coerce,
            force_responsible_sales=force_responsible_sales,
        )

        # 没有任何有效字段值的行跳过
        if not pr.raw_data and not pr.parsed_data:
            continue

        rows.append(pr)

    # 管理员/组长上传且 Excel 未指定业务员时，默认归到上传账号本人名下
    if default_responsible_sales:
        for row in rows:
            row.parsed_data.setdefault("responsible_sales", default_responsible_sales)

    # 正式模板：从 sheet 标题提取国家，scope_user 补充组别
    if is_formal:
        country = _extract_country_from_sheet(ws)
        for row in rows:
            _fill_formal_defaults(row, scope_user, country)

    # 自动填入产品大类（product_category 为空时根据品名推断）
    for row in rows:
        if not row.parsed_data.get("product_category"):
            detected = detect_product_category(row.parsed_data.get("product_name"))
            if detected:
                row.parsed_data["product_category"] = detected

    return ParseResult(
        file_name=file_name,
        sheet_name=sheet_name,
        rows=rows,
        column_mapping=column_mapping,
        missing_headers=missing_headers,
        unmapped_headers=unmapped_headers,
    )
