"""
询单总表 Excel 导出服务

用 openpyxl 生成格式化 xlsx：
  - 中文表头 + 蓝色加粗
  - 冻结首行 + 自动筛选
  - 日期格式 YYYY-MM-DD
  - 金额保留两位小数
  - 毛利润率显示为 "18.50%" 形式
  - 根据预设宽度调整列宽
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Inquiry

# (中文表头, model 字段名, 列宽)
EXPORT_COLUMNS: list[tuple[str, str, int]] = [
    ("询单号",      "inquiry_no",          18),
    ("客户代码",    "customer_code",        12),
    ("客户订单号",  "customer_order_no",    16),
    ("客户名称",    "customer_name",        20),
    ("客户简称",    "customer_short_name",  12),
    ("国家",        "country",              10),
    ("地区",        "region",               10),
    ("客户类别",    "customer_category",    12),
    ("所属小组",    "group_name",           10),
    ("负责业务员",  "responsible_sales",    12),
    ("协助业务员",  "assisting_sales",      12),
    ("产品大类",    "product_category",     10),
    ("品名",        "product_name",         26),
    ("系列",        "series_name",          20),
    ("季节",        "season",               12),
    ("数量",        "quantity",              8),
    ("询单日期",    "inquiry_date",         13),
    ("报价情况",    "quote_status",         10),
    ("订单状态",    "order_status",         10),
    ("最终报价",    "final_quote",          13),
    ("工厂价格",    "factory_price",        13),
    ("毛利润率",    "gross_profit_rate",    10),
    ("下单单价",    "order_unit_price",     13),
    ("下单数量",    "order_quantity",        8),
    ("贸易额",      "trade_amount",         14),
    ("下单日期",    "order_date",           13),
    ("询单年份",    "inquiry_year",         10),
    ("询单月份",    "inquiry_month",        10),
    ("备注",        "remark",               28),
]

_DATE_FIELDS   = {"inquiry_date", "order_date"}
_AMOUNT_FIELDS = {"final_quote", "factory_price", "order_unit_price", "trade_amount"}
_PCT_FIELDS    = {"gross_profit_rate"}


def _get(inq: Inquiry, field: str) -> Any:
    v = getattr(inq, field, None)
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.date()
    return v


def build_inquiry_excel(rows: list[Inquiry]) -> bytes:
    """
    将询单列表渲染为 xlsx bytes。
    rows 为空时返回只含表头的 Excel（符合产品要求，不报错）。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "询单总表"

    hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (header, _, width) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    date_fmt   = "YYYY-MM-DD"
    amount_fmt = "#,##0.00"
    pct_fmt    = '0.00"%"'   # 18.5 → "18.50%"，不乘以 100

    for row_idx, inq in enumerate(rows, start=2):
        for col_idx, (_, field, _) in enumerate(EXPORT_COLUMNS, start=1):
            v    = _get(inq, field)
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(vertical="center")
            if v is None:
                continue
            if field in _DATE_FIELDS:
                cell.number_format = date_fmt
            elif field in _AMOUNT_FIELDS:
                cell.number_format = amount_fmt
            elif field in _PCT_FIELDS:
                cell.number_format = pct_fmt

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选（仅表头行，避免数据为空时 dimensions 异常）
    last_col = get_column_letter(len(EXPORT_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
