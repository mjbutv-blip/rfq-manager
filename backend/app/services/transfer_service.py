"""
一键转单 Excel 生成服务
生成：工厂购销合同基础版 + 财务转单统计表基础版
"""

import os
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.models.inquiry import Inquiry

# 文件保存根目录（相对于 backend/ 目录运行位置）
_TRANSFER_DIR = Path(__file__).resolve().parent.parent.parent / "generated" / "transfers"


def _ensure_dir() -> None:
    _TRANSFER_DIR.mkdir(parents=True, exist_ok=True)


def _set_col_widths(ws) -> None:
    """根据内容自动调整列宽，最小 10，最大 40。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # CJK 字符计 2 宽
                s = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in s)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def _header_font() -> Font:
    return Font(bold=True, size=12)


def _field_font() -> Font:
    return Font(bold=True, size=10)


def _title_fill() -> PatternFill:
    return PatternFill("solid", fgColor="D9E1F2")


def _fmt_date(d) -> str | None:
    if d is None:
        return None
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


def _safe_float(v) -> float | None:
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


# ── 工厂购销合同基础版 ─────────────────────────────────────────────────────────

_FACTORY_HEADERS = [
    ("询单号",       "inquiry_no"),
    ("客户订单号",    "customer_order_no"),
    ("客户名称",      "customer_name"),
    ("客户简称",      "customer_short_name"),
    ("所属小组",      "group_name"),
    ("负责业务员",    "responsible_sales"),
    ("产品大类",      "product_category"),
    ("品名",          "product_name"),
    ("系列",          "series_name"),
    ("季节",          "season"),
    ("订单数量",      "order_quantity"),
    ("下单单价(USD)", "order_unit_price"),
    ("贸易额(USD)",   "trade_amount"),
    ("工厂价格(CNY)", "factory_price"),
    ("下单日期",      "order_date"),
    ("备注",          "remark"),
    ("生产要求",      None),
    ("验货标准",      None),
]


def generate_factory_contract(inquiry: Inquiry, ts: str) -> Path:
    """生成工厂购销合同 Excel，返回文件路径。"""
    _ensure_dir()
    filename = f"工厂购销合同_{inquiry.inquiry_no}_{ts}.xlsx"
    filepath = _TRANSFER_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "工厂购销合同"

    # 第一行：标题
    ws.merge_cells(f"A1:{get_column_letter(len(_FACTORY_HEADERS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"工厂购销合同（基础版）—— {inquiry.inquiry_no}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = _title_fill()
    ws.row_dimensions[1].height = 28

    # 第二行：字段名
    for col_idx, (header, _) in enumerate(_FACTORY_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _field_font()
        cell.fill = PatternFill("solid", fgColor="EBF1DE")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 第三行：数据
    for col_idx, (_, field) in enumerate(_FACTORY_HEADERS, start=1):
        if field is None:
            ws.cell(row=3, column=col_idx, value="待补充")
            continue
        raw = getattr(inquiry, field, None)
        if field == "order_date":
            ws.cell(row=3, column=col_idx, value=_fmt_date(raw))
        elif field in ("order_unit_price", "trade_amount", "factory_price"):
            fv = _safe_float(raw)
            c = ws.cell(row=3, column=col_idx, value=fv)
            if fv is not None:
                c.number_format = "#,##0.00"
        else:
            ws.cell(row=3, column=col_idx, value=raw)
    ws.row_dimensions[3].height = 18

    _set_col_widths(ws)
    wb.save(filepath)
    return filepath


# ── 财务转单统计表基础版 ───────────────────────────────────────────────────────

_FINANCE_HEADERS = [
    ("询单号",        "inquiry_no"),
    ("客户订单号",    "customer_order_no"),
    ("客户名称",      "customer_name"),
    ("客户简称",      "customer_short_name"),
    ("所属小组",      "group_name"),
    ("负责业务员",    "responsible_sales"),
    ("订单状态",      "order_status"),
    ("报价情况",      "quote_status"),
    ("最终报价(USD)", "final_quote"),
    ("工厂价格(CNY)", "factory_price"),
    ("下单单价(USD)", "order_unit_price"),
    ("下单数量",      "order_quantity"),
    ("贸易额(USD)",   "trade_amount"),
    ("毛利润率",      "gross_profit_rate"),
    ("毛利润额(USD)", "__gross_profit_amount"),   # 计算字段
    ("下单日期",      "order_date"),
    ("备注",          "remark"),
]

_MONEY_FIELDS = {"final_quote", "factory_price", "order_unit_price", "trade_amount"}


def _calc_gross_profit_amount(inq: Inquiry) -> float | None:
    """毛利润额 = 贸易额 - 工厂价格 × 下单数量"""
    trade  = _safe_float(inq.trade_amount)
    fp     = _safe_float(inq.factory_price)
    qty    = inq.order_quantity
    if trade is not None and fp is not None and qty is not None:
        return round(trade - fp * qty, 2)
    return None


def generate_finance_transfer(inquiry: Inquiry, ts: str) -> Path:
    """生成财务转单统计表 Excel，返回文件路径。"""
    _ensure_dir()
    filename = f"财务转单统计表_{inquiry.inquiry_no}_{ts}.xlsx"
    filepath = _TRANSFER_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "财务转单统计表"

    # 标题行
    ws.merge_cells(f"A1:{get_column_letter(len(_FINANCE_HEADERS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"财务转单统计表（基础版）—— {inquiry.inquiry_no}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = _title_fill()
    ws.row_dimensions[1].height = 28

    # 字段名行
    for col_idx, (header, _) in enumerate(_FINANCE_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _field_font()
        cell.fill = PatternFill("solid", fgColor="EBF1DE")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 数据行
    for col_idx, (_, field) in enumerate(_FINANCE_HEADERS, start=1):
        if field == "__gross_profit_amount":
            fv = _calc_gross_profit_amount(inquiry)
            c = ws.cell(row=3, column=col_idx, value=fv)
            if fv is not None:
                c.number_format = "#,##0.00"
            continue
        if field is None:
            continue

        raw = getattr(inquiry, field, None)

        if field == "order_date":
            ws.cell(row=3, column=col_idx, value=_fmt_date(raw))
        elif field == "gross_profit_rate":
            fv = _safe_float(raw)
            c = ws.cell(row=3, column=col_idx, value=fv / 100 if fv is not None else None)
            if fv is not None:
                c.number_format = "0.00%"
        elif field in _MONEY_FIELDS:
            fv = _safe_float(raw)
            c = ws.cell(row=3, column=col_idx, value=fv)
            if fv is not None:
                c.number_format = "#,##0.00"
        else:
            ws.cell(row=3, column=col_idx, value=raw)
    ws.row_dimensions[3].height = 18

    _set_col_widths(ws)
    wb.save(filepath)
    return filepath


# ── 主入口 ─────────────────────────────────────────────────────────────────────

def detect_missing_fields(inquiry: Inquiry) -> list[str]:
    """检查关键字段是否缺失，返回缺失字段中文名列表。"""
    checks = [
        (inquiry.order_quantity,   "下单数量"),
        (inquiry.order_unit_price, "下单单价"),
        (inquiry.trade_amount,     "贸易额"),
        (inquiry.factory_price,    "工厂价格"),
        (inquiry.customer_short_name or inquiry.customer_name, "客户信息"),
    ]
    return [label for value, label in checks if value is None]


def generate_transfer_files(inquiry: Inquiry) -> tuple[Path, Path, str]:
    """
    生成两个 Excel 文件。
    返回 (factory_path, finance_path, ts_str)。
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    factory_path = generate_factory_contract(inquiry, ts)
    finance_path = generate_finance_transfer(inquiry, ts)
    return factory_path, finance_path, ts
