from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import openpyxl
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app import crud
from app.core.permissions import can_edit_inquiry
from app.models import Customer, ImportBatch, ImportRow, Inquiry, InquiryItem, OperationLog
from app.services.operation_log_service import log_kwargs_from_user, safe_log

SOURCE_SHEETS = ("总表", "总表海外", "海外报价表-美金")


@dataclass
class BaseImportRow:
    source_sheet: str
    row_number: int
    inquiry_no: str | None
    customer_order_no: str | None = None
    season: str | None = None
    order_status: str | None = None
    inquiry_date: date | None = None
    customer_code: str | None = None
    product_name: str | None = None
    product_category: str | None = None
    series_name: str | None = None
    quantity: int | None = None
    style_no: str | None = None
    notes: str | None = None
    raw_data: dict[str, Any] | None = None


def _clean_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in {"#N/A", "#DIV/0!", "#NUM!", "#VALUE!"}:
        return None
    return s


def _clean_optional(v: Any) -> str | None:
    s = _clean_str(v)
    if not s:
        return None
    if s in {"0", "0.0", "—", "-", "--", "——", "———", "————", "_", "__", "___", "/", "／"}:
        return None
    return s


def _clean_inquiry_no(v: Any) -> str | None:
    s = _clean_str(v)
    if not s:
        return None
    first = re.split(r"[\s（(]", s, maxsplit=1)[0].strip()
    if not re.search(r"[A-Za-z]", first):
        return None
    if not re.match(r"^[A-Za-z0-9][A-Za-z0-9_-]*$", first):
        return None
    return first


def _to_int(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v)
    total_match = re.search(r"共\s*(\d[\d,]*)\s*件", s)
    if total_match:
        return int(total_match.group(1).replace(",", ""))
    matches = [int(m.replace(",", "")) for m in re.findall(r"\d[\d,]*", s)]
    if not matches:
        return None
    return max(matches)


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def _json_value(v: Any) -> Any:
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, uuid.UUID):
        return str(v)
    return v


def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return not v.strip()
    return False


def _fill_empty_inquiry_fields(inquiry: Inquiry, row: BaseImportRow, customer: Customer | None = None) -> dict[str, Any]:
    updates: dict[str, Any] = {}
    candidates = (
        ("customer_code", row.customer_code),
        ("customer_short_name", customer.customer_short_name if customer else None),
        ("customer_order_no", row.customer_order_no),
        ("season", row.season),
        ("order_status", row.order_status),
        ("inquiry_date", row.inquiry_date),
        ("product_category", row.product_category),
        ("product_name", row.product_name),
        ("series_name", row.series_name),
        ("quantity", row.quantity),
        ("remark", row.notes),
    )
    for field, value in candidates:
        if value is not None and _is_empty(getattr(inquiry, field, None)):
            setattr(inquiry, field, value)
            updates[field] = _json_value(value)
    return updates


def _header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    max_header_row = min(ws.max_row, 4)
    for r in range(1, max_header_row + 1):
        for c in range(1, ws.max_column + 1):
            val = _clean_str(ws.cell(r, c).value)
            if val and val not in headers:
                headers[val] = c
    return headers


def _find_header(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    for header, col in headers.items():
        compact = header.replace("\n", "").replace(" ", "")
        if any(name in compact for name in names):
            return col
    return None


def _cell(ws, row: int, col: int | None) -> Any:
    return ws.cell(row, col).value if col else None


def _notes(*parts: tuple[str, Any]) -> str | None:
    chunks = []
    for label, value in parts:
        s = _clean_str(value)
        if s:
            chunks.append(f"{label}：{s}")
    return "\n\n".join(chunks) if chunks else None


def _item_key(row: BaseImportRow) -> tuple[str, str | None, bool]:
    if row.style_no:
        return ("style", row.style_no.strip().lower(), False)
    product = (row.product_name or "").strip().lower()
    series = (row.series_name or "").strip().lower()
    if product or series:
        return ("product_series", f"{product}|{series}", False)
    return ("uncertain", None, True)


def _parse_workbook(file_bytes: bytes, uniform_customer_code: str | None = None) -> tuple[list[BaseImportRow], dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows: list[BaseImportRow] = []
    sheet_stats: dict[str, Any] = {}
    uniform_customer_code = _clean_optional(uniform_customer_code)

    for sheet_name in SOURCE_SHEETS:
        if sheet_name not in wb.sheetnames:
            sheet_stats[sheet_name] = {"rows": 0, "present": False}
            continue
        ws = wb[sheet_name]
        headers = _header_map(ws)
        inquiry_col = _find_header(headers, "询单号")
        order_col = _find_header(headers, "订单号", "客户订单号")
        series_col = _find_header(headers, "系列")
        season_col = _find_header(headers, "季节")
        product_col = _find_header(headers, "品名")
        qty_col = _find_header(headers, "数量") if sheet_name in {"总表", "总表海外"} else None
        style_col = _find_header(headers, "做工图/尺码表", "款号", "style_no")
        order_status_col = 20 if sheet_name in {"总表", "总表海外"} else None
        customer_code_col = _find_header(headers, "客户代码", "客户编号")
        inquiry_date_col = _find_header(headers, "询单日期")

        start_row = 3 if sheet_name in {"总表", "总表海外"} else 5
        parsed_rows = 0
        for r in range(start_row, ws.max_row + 1):
            raw_inquiry_no = _clean_str(_cell(ws, r, inquiry_col))
            inquiry_no = _clean_inquiry_no(raw_inquiry_no)
            customer_order_no = _clean_optional(_cell(ws, r, order_col))
            product_name = _clean_optional(_cell(ws, r, product_col))
            if not inquiry_no and not any(_clean_str(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 8) + 1)):
                continue
            if not inquiry_no and raw_inquiry_no and not re.search(r"[A-Za-z]", raw_inquiry_no):
                continue
            if not inquiry_no and not customer_order_no and not product_name:
                continue
            if not inquiry_no:
                rows.append(BaseImportRow(
                    source_sheet=sheet_name,
                    row_number=r,
                    inquiry_no=None,
                    raw_data={"error": "询单号为空", "source": f"{sheet_name}!{r}"},
                ))
                parsed_rows += 1
                continue

            if not customer_order_no and not product_name:
                continue

            customer_code = _clean_optional(_cell(ws, r, customer_code_col)) or uniform_customer_code
            row = BaseImportRow(
                source_sheet=sheet_name,
                row_number=r,
                inquiry_no=inquiry_no,
                customer_order_no=customer_order_no,
                season=_clean_optional(_cell(ws, r, season_col)),
                order_status=_clean_optional(_cell(ws, r, order_status_col)),
                inquiry_date=_to_date(_cell(ws, r, inquiry_date_col)),
                customer_code=customer_code,
                product_name=product_name,
                product_category=None,
                series_name=_clean_optional(_cell(ws, r, series_col)),
                quantity=_to_int(_cell(ws, r, qty_col)),
                style_no=_clean_optional(_cell(ws, r, style_col)),
                notes=_notes(
                    ("强调内容", ws.cell(r, 7).value if sheet_name in {"总表", "总表海外"} else None),
                    ("建议内容", ws.cell(r, 8).value if sheet_name in {"总表", "总表海外"} else None),
                    ("工厂回复", ws.cell(r, 9).value if sheet_name in {"总表", "总表海外"} else None),
                ),
                raw_data={
                    "source_sheet": sheet_name,
                    "row_number": r,
                    "inquiry_no": inquiry_no,
                    "customer_code_source": "excel" if _clean_optional(_cell(ws, r, customer_code_col)) else ("uniform_input" if customer_code else None),
                },
            )
            rows.append(row)
            parsed_rows += 1
        sheet_stats[sheet_name] = {
            "rows": parsed_rows,
            "present": True,
            "layout": "一行一个款式/报价场景，询单号可跨多行",
            "has_customer_code_column": customer_code_col is not None,
        }
    return rows, sheet_stats


async def _load_customers(db: AsyncSession, codes: set[str]) -> dict[str, Customer]:
    if not codes:
        return {}
    result = await db.execute(select(Customer).where(Customer.customer_code.in_(codes)))
    return {c.customer_code: c for c in result.scalars().all()}


async def _load_inquiries(db: AsyncSession, inquiry_nos: set[str]) -> dict[str, Inquiry]:
    if not inquiry_nos:
        return {}
    result = await db.execute(select(Inquiry).where(Inquiry.inquiry_no.in_(inquiry_nos)))
    return {i.inquiry_no: i for i in result.scalars().all()}


async def _load_items(db: AsyncSession, inquiry_nos: set[str]) -> dict[str, set[tuple[str, str | None]]]:
    if not inquiry_nos:
        return {}
    result = await db.execute(select(InquiryItem).where(InquiryItem.inquiry_no.in_(inquiry_nos)))
    items: dict[str, set[tuple[str, str | None]]] = {}
    for item in result.scalars().all():
        pseudo = BaseImportRow(
            source_sheet="db",
            row_number=0,
            inquiry_no=item.inquiry_no,
            product_name=item.product_name,
            series_name=item.series_name,
            style_no=item.style_no,
        )
        key_type, key, uncertain = _item_key(pseudo)
        if not uncertain:
            items.setdefault(item.inquiry_no or "", set()).add((key_type, key))
    return items


def _row_payload(row: BaseImportRow) -> dict[str, Any]:
    return {
        "source_sheet": row.source_sheet,
        "row_number": row.row_number,
        "inquiry_no": row.inquiry_no,
        "customer_order_no": row.customer_order_no,
        "season": row.season,
        "order_status": row.order_status,
        "inquiry_date": _json_value(row.inquiry_date),
        "customer_code": row.customer_code,
        "product_name": row.product_name,
        "product_category": row.product_category,
        "series_name": row.series_name,
        "quantity": row.quantity,
        "style_no": row.style_no,
        "notes": row.notes,
    }


async def preview_base_inquiry_import(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str,
    user: Any,
    uniform_customer_code: str | None = None,
) -> dict[str, Any]:
    parsed_rows, sheet_stats = _parse_workbook(file_bytes, uniform_customer_code)
    inquiry_nos = {r.inquiry_no for r in parsed_rows if r.inquiry_no}
    customer_codes = {r.customer_code for r in parsed_rows if r.customer_code}
    inquiries = await _load_inquiries(db, inquiry_nos)
    customers = await _load_customers(db, customer_codes)
    existing_items = await _load_items(db, inquiry_nos)
    seen_items: dict[str, set[tuple[str, str | None]]] = {}
    seen_new_inquiries: set[str] = set()

    summary = {
        "total_rows": 0,
        "new_inquiries": 0,
        "new_items": 0,
        "existing_inquiries": 0,
        "duplicate_items": 0,
        "customer_unmatched": 0,
        "item_identity_uncertain": 0,
        "failed": 0,
        "importable_rows": 0,
    }
    rows: list[dict[str, Any]] = []

    for row in parsed_rows:
        summary["total_rows"] += 1
        flags: list[str] = []
        errors: list[str] = []
        status = "new_inquiry"
        if not row.inquiry_no:
            status = "failed"
            errors.append("Excel 询单号为空")
            summary["failed"] += 1
            rows.append({
                **_row_payload(row),
                "status": status,
                "flags": flags,
                "errors": errors,
                "item_identity_key": None,
                "customer_matched": False,
                "customer_will_create": False,
                "can_confirm": False,
            })
            continue

        existing_inquiry = inquiries.get(row.inquiry_no)
        if existing_inquiry:
            status = "existing_inquiry"
            flags.append("existing_inquiry")
            summary["existing_inquiries"] += 1
            can_write_inquiry = can_edit_inquiry(existing_inquiry, user)
            if not can_write_inquiry:
                status = "failed"
                errors.append("无权限向该询单追加款式")
        elif row.inquiry_no in seen_new_inquiries:
            status = "new_item_for_existing_inquiry"
        else:
            status = "new_inquiry"
            seen_new_inquiries.add(row.inquiry_no)
            summary["new_inquiries"] += 1

        if not row.customer_code:
            flags.append("customer_unmatched")
            summary["customer_unmatched"] += 1
        customer_will_create = bool(row.customer_code and row.customer_code not in customers)

        key_type, key, uncertain = _item_key(row)
        if uncertain:
            flags.append("item_identity_uncertain")
            summary["item_identity_uncertain"] += 1
            duplicate = False
        else:
            key_tuple = (key_type, key)
            duplicate = key_tuple in existing_items.get(row.inquiry_no, set()) or key_tuple in seen_items.get(row.inquiry_no, set())
        if duplicate:
            status = "duplicate_item"
            flags.append("duplicate_item")
            summary["duplicate_items"] += 1
        else:
            seen_items.setdefault(row.inquiry_no, set()).add((key_type, key))
            if status != "failed":
                summary["new_items"] += 1

        can_confirm = status not in {"failed", "duplicate_item"}
        if can_confirm:
            summary["importable_rows"] += 1

        rows.append({
            **_row_payload(row),
            "status": status,
            "flags": flags,
            "errors": errors,
            "item_identity_key": f"{key_type}:{key}" if key else key_type,
            "customer_matched": bool(row.customer_code and row.customer_code in customers),
            "customer_will_create": customer_will_create,
            "can_confirm": can_confirm,
        })

    return {
        "file_name": file_name,
        "sheet_stats": sheet_stats,
        "summary": summary,
        "rows": rows,
        "uniform_customer_code": _clean_optional(uniform_customer_code),
    }


async def _log_in_session(db: AsyncSession, user: Any, **kwargs: Any) -> None:
    if "before_data" in kwargs:
        kwargs["before_data_json"] = kwargs.pop("before_data")
    if "after_data" in kwargs:
        kwargs["after_data_json"] = kwargs.pop("after_data")
    log = OperationLog(**log_kwargs_from_user(user), **kwargs)
    db.add(log)


async def _get_or_create_customer(db: AsyncSession, customer_code: str | None) -> tuple[Customer | None, bool]:
    if not customer_code:
        return None, False
    customer = (await db.execute(select(Customer).where(Customer.customer_code == customer_code))).scalars().first()
    if customer:
        return customer, False
    customer = Customer(customer_code=customer_code)
    db.add(customer)
    await db.flush()
    return customer, True


async def confirm_base_inquiry_import(
    db: AsyncSession,
    file_bytes: bytes,
    file_name: str,
    user: Any,
    uniform_customer_code: str | None = None,
) -> dict[str, Any]:
    preview = await preview_base_inquiry_import(db, file_bytes, file_name, user, uniform_customer_code)
    batch = await crud.create_import_batch(db, {
        "file_name": file_name,
        "uploaded_by": getattr(user, "username", None),
        "total_rows": preview["summary"]["total_rows"],
        "status": "pending",
    })
    await db.flush()

    summary = {
        "created_inquiries": 0,
        "created_items": 0,
        "updated_inquiry_fields": 0,
        "existing_inquiries": 0,
        "duplicate_items_skipped": 0,
        "customer_records_created": 0,
        "customer_unmatched_rows": 0,
        "uncertain_item_rows": 0,
        "write_failed_rows": 0,
    }
    results: list[dict[str, Any]] = []
    created_inquiries: dict[str, Inquiry] = {}
    created_item_keys: dict[str, set[tuple[str, str | None]]] = {}

    for row_data in preview["rows"]:
        status = row_data["status"]
        row = BaseImportRow(
            source_sheet=row_data["source_sheet"],
            row_number=row_data["row_number"],
            inquiry_no=row_data["inquiry_no"],
            customer_order_no=row_data["customer_order_no"],
            season=row_data["season"],
            order_status=row_data["order_status"],
            inquiry_date=_to_date(row_data["inquiry_date"]),
            customer_code=row_data["customer_code"],
            product_name=row_data["product_name"],
            product_category=row_data["product_category"],
            series_name=row_data["series_name"],
            quantity=row_data["quantity"],
            style_no=row_data["style_no"],
            notes=row_data["notes"],
        )
        import_status = status
        error_message = None

        if status == "failed":
            summary["write_failed_rows"] += 1
        else:
            try:
                async with db.begin_nested():
                    customer, customer_created = await _get_or_create_customer(db, row.customer_code)
                    if customer_created:
                        summary["customer_records_created"] += 1

                    inquiry = created_inquiries.get(row.inquiry_no or "")
                    if inquiry is None:
                        inquiry = (await db.execute(select(Inquiry).where(Inquiry.inquiry_no == row.inquiry_no))).scalars().first()

                    if inquiry:
                        if row_data["status"] == "existing_inquiry":
                            summary["existing_inquiries"] += 1
                        if not can_edit_inquiry(inquiry, user):
                            raise PermissionError("无权限向该询单追加款式")
                        updates = _fill_empty_inquiry_fields(inquiry, row, customer)
                        if updates:
                            summary["updated_inquiry_fields"] += len(updates)
                            await _log_in_session(
                                db,
                                user,
                                action_type="base_inquiry_import_field_fill_from_excel",
                                target_type="inquiry",
                                target_id=str(inquiry.id),
                                inquiry_id=inquiry.id,
                                inquiry_no=inquiry.inquiry_no,
                                description="基础询单导入补齐询单主表空字段",
                                after_data={
                                    "excel_file": file_name,
                                    "sheet": row.source_sheet,
                                    "row": row.row_number,
                                    "updated_fields": updates,
                                },
                            )
                    else:
                        inquiry = Inquiry(
                            inquiry_no=row.inquiry_no or "",
                            customer_code=row.customer_code,
                            customer_short_name=customer.customer_short_name if customer else None,
                            customer_order_no=row.customer_order_no,
                            season=row.season,
                            order_status=row.order_status,
                            inquiry_date=row.inquiry_date,
                            product_category=row.product_category,
                            product_name=row.product_name,
                            series_name=row.series_name,
                            quantity=row.quantity,
                            group_name=getattr(user, "group_name", None) if getattr(user, "role", None) == "group_leader" else None,
                            responsible_sales=None,
                            remark=row.notes,
                            import_batch_id=batch.id,
                        )
                        db.add(inquiry)
                        await db.flush()
                        created_inquiries[row.inquiry_no or ""] = inquiry
                        summary["created_inquiries"] += 1
                        await _log_in_session(
                            db,
                            user,
                            action_type="base_inquiry_create_from_excel",
                            target_type="inquiry",
                            target_id=str(inquiry.id),
                            inquiry_id=inquiry.id,
                            inquiry_no=inquiry.inquiry_no,
                            description="从 Excel 创建基础询单",
                            after_data={**_row_payload(row), "excel_file": file_name},
                        )

                    if status == "duplicate_item":
                        summary["duplicate_items_skipped"] += 1
                        import_status = "skipped_duplicate_item"
                        if not row.customer_code:
                            summary["customer_unmatched_rows"] += 1
                        skip_item = True
                    else:
                        skip_item = False

                    if not skip_item:
                        key_type, key, uncertain = _item_key(row)
                        if uncertain:
                            summary["uncertain_item_rows"] += 1
                        else:
                            current_keys = created_item_keys.setdefault(row.inquiry_no or "", set())
                            if (key_type, key) in current_keys:
                                summary["duplicate_items_skipped"] += 1
                                import_status = "skipped_duplicate_item"
                                skip_item = True
                            else:
                                current_keys.add((key_type, key))
                    else:
                        key_type, key, uncertain = _item_key(row)

                    if not skip_item:
                        item = InquiryItem(
                            inquiry_id=inquiry.id,
                            inquiry_no=inquiry.inquiry_no,
                            product_name=row.product_name,
                            product_category=row.product_category,
                            series_name=row.series_name,
                            quantity=row.quantity,
                            style_no=row.style_no,
                            order_status=row.order_status,
                            remark=f"来源：{row.source_sheet}!{row.row_number}",
                            extra_data={
                                "source_file": file_name,
                                "source_sheet": row.source_sheet,
                                "source_row": row.row_number,
                                "item_identity_key": row_data.get("item_identity_key"),
                                "item_identity_uncertain": uncertain,
                            },
                        )
                        db.add(item)
                        await db.flush()
                        summary["created_items"] += 1
                        if not row.customer_code:
                            summary["customer_unmatched_rows"] += 1
                            await _log_in_session(
                                db,
                                user,
                                action_type="base_inquiry_customer_unmatched",
                                target_type="inquiry",
                                target_id=str(inquiry.id),
                                inquiry_id=inquiry.id,
                                inquiry_no=inquiry.inquiry_no,
                                description="基础询单导入时客户未匹配",
                                after_data={"excel_file": file_name, "sheet": row.source_sheet, "row": row.row_number},
                            )
                        await _log_in_session(
                            db,
                            user,
                            action_type="base_inquiry_item_create_from_excel",
                            target_type="inquiry_item",
                            target_id=str(item.id),
                            inquiry_id=inquiry.id,
                            inquiry_no=inquiry.inquiry_no,
                            description="从 Excel 创建询单款式明细",
                            after_data={**_row_payload(row), "item_identity_key": row_data.get("item_identity_key"), "excel_file": file_name},
                        )
                        import_status = "imported"
            except Exception as exc:
                summary["write_failed_rows"] += 1
                error_message = str(exc)
                import_status = "error"

        db.add(ImportRow(
            batch_id=batch.id,
            row_number=row.row_number,
            inquiry_no=row.inquiry_no,
            status=import_status,
            raw_data_json={"file_name": file_name, "sheet": row.source_sheet, "row_number": row.row_number},
            parsed_data_json=_row_payload(row),
            error_message=error_message,
        ))
        results.append({**row_data, "result_status": import_status, "error_message": error_message})

    batch.success_rows = summary["created_items"]
    batch.failed_rows = summary["write_failed_rows"]
    batch.new_rows = summary["created_inquiries"]
    batch.existing_rows = summary["existing_inquiries"]
    batch.duplicate_rows = summary["duplicate_items_skipped"]
    batch.uncertain_rows = summary["uncertain_item_rows"]
    batch.validation_failed_rows = preview["summary"]["failed"]
    batch.write_failed_rows = summary["write_failed_rows"]
    batch.status = "success" if summary["write_failed_rows"] == 0 else "partial"

    await safe_log(
        **log_kwargs_from_user(user),
        action_type="base_inquiry_import_confirm",
        target_type="import_batch",
        target_id=batch.id,
        description="确认基础询单 Excel 导入",
        after_data={"file_name": file_name, "summary": summary},
    )

    return {
        "file_name": file_name,
        "batch_id": str(batch.id),
        "summary": summary,
        "rows": results,
        "next_step": {
            "message": "基础询单创建完成。下一步请进入“来龙去脉表资料导入”，回填报价轮次、工厂报价和订单资料。",
            "path": "/inquiry-journey-import",
        },
    }
